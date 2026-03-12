import os, io, math, json, base64
from flask import Flask, request, jsonify, send_file, render_template
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024

# ── CONSTANTS ──────────────────────────────────────────────────────────────────
START_MIN     = 8 * 60
LUNCH_START   = 13 * 60
LUNCH_END     = 13 * 60 + 45
LUNCH_DUR     = 45
EFFECTIVE_MIN = 9 * 60 - LUNCH_DUR   # 495 min

def fmt_time(m):
    t = int(round(m))
    return f"{t//60:02d}:{t%60:02d}"

def advance_clock(clock, dur):
    if clock >= LUNCH_START and clock < LUNCH_END:
        clock = LUNCH_END
    end = clock + dur
    if clock < LUNCH_START < end:
        end += LUNCH_DUR
    return end

def run_picker_plan(file_bytes):
    xl = pd.read_excel(io.BytesIO(file_bytes), sheet_name=None)
    do_df      = xl['DO'].copy()
    machine_df = xl['MACHINE'].copy()
    machine_df.rename(columns={'SCANNER NAME': 'SCANNER_NAME'}, inplace=True, errors='ignore')
    if 'SCANNER_NAME' not in machine_df.columns and 'SCANNER NAME' in machine_df.columns:
        machine_df.rename(columns={'SCANNER NAME': 'SCANNER_NAME'}, inplace=True)

    BGT_COL = 'BGT/PICKER'
    bgt_val  = int(machine_df[BGT_COL].mode()[0])
    items_per_min = bgt_val / EFFECTIVE_MIN

    agg = (do_df.groupby(['FLOOR', 'SEC'], sort=True)['DO_QTY']
           .sum().reset_index().sort_values(['FLOOR', 'SEC']))

    total_qty   = int(do_df['DO_QTY'].sum())
    req_pickers = math.ceil(total_qty / bgt_val)

    machines = machine_df.reset_index(drop=True)
    m_idx, plans = 0, []

    for floor in sorted(agg['FLOOR'].unique()):
        if m_idx >= len(machines):
            break
        floor_data = agg[agg['FLOOR'] == floor]
        cur_mach   = machines.iloc[m_idx]
        cap_used   = 0
        clock      = START_MIN

        for _, sec_row in floor_data.iterrows():
            sec      = sec_row['SEC']
            qty_left = int(sec_row['DO_QTY'])

            while qty_left > 0:
                available = bgt_val - cap_used
                if available == 0:
                    m_idx += 1
                    if m_idx >= len(machines):
                        break
                    cur_mach  = machines.iloc[m_idx]
                    cap_used  = 0
                    clock     = START_MIN
                    available = bgt_val

                assign     = min(qty_left, available)
                dur        = assign / items_per_min
                task_start = LUNCH_END if (clock >= LUNCH_START and clock < LUNCH_END) else clock
                task_end   = advance_clock(task_start, dur)
                cap_used  += assign
                clock      = task_end

                plans.append({
                    'MACHINE_NO':            cur_mach['MACHINE_NO'],
                    'SCANNER_NAME':          cur_mach['SCANNER_NAME'],
                    'GROUP':                 cur_mach['GROUP'],
                    'FLOOR':                 int(floor),
                    'SEC':                   sec,
                    'TOTAL_DO_QTY_ASSIGNED': assign,
                    'START_TIME':            fmt_time(task_start),
                    'END_TIME':              fmt_time(task_end),
                    'UTILIZATION_%':         round(cap_used / bgt_val * 100, 1),
                    'REMAINING_CAPACITY':    bgt_val - cap_used,
                })
                qty_left -= assign
        m_idx += 1

    plan_df = pd.DataFrame(plans)

    # ── FLOOR SUMMARY ──────────────────────────────────────────────────────────
    floor_summary = (do_df.groupby('FLOOR')['DO_QTY'].sum()
                     .reset_index().rename(columns={'DO_QTY': 'TOTAL_QTY'}))
    floor_summary['REQ_PICKERS'] = (floor_summary['TOTAL_QTY'] / bgt_val).apply(math.ceil)
    assigned_pf  = (plan_df.groupby('FLOOR')['MACHINE_NO'].nunique()
                    .reset_index().rename(columns={'MACHINE_NO': 'ASSIGNED'}))
    avg_util_pf  = (plan_df.groupby('FLOOR')['UTILIZATION_%'].mean()
                    .reset_index().rename(columns={'UTILIZATION_%': 'AVG_UTIL'}))
    floor_merged = (floor_summary.merge(assigned_pf, on='FLOOR', how='left')
                    .merge(avg_util_pf, on='FLOOR', how='left').fillna(0))

    # ── BUILD EXCEL ────────────────────────────────────────────────────────────
    wb = load_workbook(io.BytesIO(file_bytes))
    for name in ['PICKER_PLAN', 'SUMMARY']:
        if name in wb.sheetnames:
            del wb[name]

    def fill(h): return PatternFill('solid', start_color=h, end_color=h)
    def brd():
        s = Side(style='thin', color='FFBDC3C7')
        return Border(left=s, right=s, top=s, bottom=s)
    def ctr(): return Alignment(horizontal='center', vertical='center')

    # PICKER_PLAN sheet
    ws = wb.create_sheet('PICKER_PLAN')
    ws.merge_cells('A1:J1')
    ws['A1'] = '🏭  PICKER PLANNING SYSTEM  —  DAILY ALLOCATION PLAN'
    ws['A1'].font      = Font(name='Arial', size=13, bold=True, color='FFFFFFFF')
    ws['A1'].fill      = fill('FF1C2833')
    ws['A1'].alignment = ctr()
    ws.row_dimensions[1].height = 28

    ws.merge_cells('A2:J2')
    ws['A2'] = (f'Total DOs: {len(do_df):,}  |  Total QTY: {total_qty:,}  |  '
                f'BGT/Picker: {bgt_val:,}  |  Required Pickers: {req_pickers}  |  '
                f'Warehouse Start: 08:00  |  Lunch: 13:00–13:45')
    ws['A2'].font      = Font(name='Arial', size=9, italic=True, color='FFFFFFFF')
    ws['A2'].fill      = fill('FF2874A6')
    ws['A2'].alignment = ctr()
    ws.row_dimensions[2].height = 18

    headers = ['MACHINE NO','SCANNER NAME','GROUP','FLOOR','SEC',
               'QTY ASSIGNED','START TIME','END TIME','UTILIZATION %','REMAINING CAP']
    widths  = [14,18,10,8,8,16,13,13,17,17]
    for c,(h,w) in enumerate(zip(headers,widths),1):
        cell = ws.cell(row=3,column=c,value=h)
        cell.font      = Font(name='Arial',size=10,bold=True,color='FFFFFFFF')
        cell.fill      = fill('FF2C3E50')
        cell.alignment = ctr()
        cell.border    = brd()
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.row_dimensions[3].height = 20

    for i, rec in plan_df.iterrows():
        r = 4 + i
        rf = 'FFAED6F1' if i%2==0 else 'FFFFFFFF'
        vals = [rec['MACHINE_NO'],rec['SCANNER_NAME'],rec['GROUP'],
                rec['FLOOR'],rec['SEC'],rec['TOTAL_DO_QTY_ASSIGNED'],
                rec['START_TIME'],rec['END_TIME'],rec['UTILIZATION_%'],rec['REMAINING_CAPACITY']]
        for c,v in enumerate(vals,1):
            cell = ws.cell(row=r,column=c,value=v)
            cell.border    = brd()
            cell.font      = Font(name='Arial',size=9)
            cell.alignment = ctr()
            cell.fill      = fill(rf)
        u = rec['UTILIZATION_%']
        col = 'FF1E8449' if u>=95 else ('FFF39C12' if u>=70 else 'FFC0392B')
        ws.cell(row=r,column=9).font = Font(name='Arial',size=9,bold=True,color=col)

    ws.freeze_panes = 'A4'
    ws.auto_filter.ref = f'A3:J{3+len(plan_df)}'
    ws.sheet_view.showGridLines = False

    # SUMMARY sheet
    ss = wb.create_sheet('SUMMARY', 0)
    ss.merge_cells('A1:F1')
    ss['A1'] = '📊  PICKER PLAN — DAILY SUMMARY DASHBOARD'
    ss['A1'].font      = Font(name='Arial',size=14,bold=True,color='FFFFFFFF')
    ss['A1'].fill      = fill('FF1C2833')
    ss['A1'].alignment = ctr()
    ss.row_dimensions[1].height = 30

    kpis = [('Total DO Lines',f'{len(do_df):,}'),('Total QTY',f'{total_qty:,}'),
            ('BGT/Picker',f'{bgt_val:,}'),('Pickers Required',f'{req_pickers}'),
            ('Available Machines',f'{len(machines)}'),('Effective Work Min',f'{EFFECTIVE_MIN}')]
    kpi_fills = ['FF1A5276','FF117A65','FF6E2FD3','FF922B21','FF1F618D','FF7D6608']
    ss.row_dimensions[3].height = 38
    ss.row_dimensions[4].height = 26
    for i,((lbl,val),bg) in enumerate(zip(kpis,kpi_fills)):
        c = i+1
        ss.column_dimensions[get_column_letter(c)].width = 20
        lc = ss.cell(row=3,column=c,value=lbl)
        lc.font=Font(name='Arial',size=8,bold=True,color='FFFFFFFF')
        lc.fill=fill(bg); lc.alignment=ctr()
        vc = ss.cell(row=4,column=c,value=val)
        vc.font=Font(name='Arial',size=16,bold=True,color='FF'+bg[2:])
        vc.fill=fill('FFF8F9FA'); vc.alignment=ctr()

    # Floor table
    fh=['FLOOR','TOTAL QTY','REQ PICKERS','ASSIGNED','AVG UTIL %']
    for c,h in enumerate(fh,1):
        cell=ss.cell(row=7,column=c,value=h)
        cell.font=Font(name='Arial',size=10,bold=True,color='FFFFFFFF')
        cell.fill=fill('FF2C3E50'); cell.alignment=ctr(); cell.border=brd()
    ss.row_dimensions[7].height=20
    for i,row in floor_merged.iterrows():
        r=8+i
        rf='FFAED6F1' if i%2==0 else 'FFFFFFFF'
        for c,v in enumerate([int(row['FLOOR']),int(row['TOTAL_QTY']),int(row['REQ_PICKERS']),
                               int(row['ASSIGNED']),round(float(row['AVG_UTIL']),1)],1):
            cell=ss.cell(row=r,column=c,value=v)
            cell.font=Font(name='Arial',size=10); cell.alignment=ctr()
            cell.border=brd(); cell.fill=fill(rf)
        ss.row_dimensions[r].height=18

    ss.sheet_view.showGridLines=False

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)

    # ── JSON summary for API ───────────────────────────────────────────────────
    summary = {
        'total_dos':        len(do_df),
        'total_qty':        total_qty,
        'bgt_per_picker':   bgt_val,
        'req_pickers':      req_pickers,
        'available_machines': len(machines),
        'pickers_used':     int(plan_df['MACHINE_NO'].nunique()),
        'plan_rows':        len(plan_df),
        'floor_breakdown':  floor_merged.to_dict('records'),
        'top_rows':         plan_df.head(20).to_dict('records'),
    }
    return out.getvalue(), summary


@app.route('/')
def index():
    return render_template('index.html')


@app.route('/process', methods=['POST'])
def process():
    if 'file' not in request.files:
        return jsonify({'error': 'No file uploaded'}), 400
    f = request.files['file']
    if not f.filename.endswith(('.xlsx', '.xls')):
        return jsonify({'error': 'Please upload an Excel file (.xlsx)'}), 400
    try:
        file_bytes = f.read()
        excel_bytes, summary = run_picker_plan(file_bytes)
        excel_b64 = base64.b64encode(excel_bytes).decode()
        return jsonify({'success': True, 'summary': summary, 'excel_b64': excel_b64})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


if __name__ == '__main__':
    app.run(debug=False, host='0.0.0.0', port=5050)
