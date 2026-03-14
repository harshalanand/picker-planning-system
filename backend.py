"""
Picker Planning System — FastAPI Backend
Run: uvicorn backend:app --host 0.0.0.0 --port 8000 --reload
"""
from fastapi import FastAPI, UploadFile, File, HTTPException, Query
from fastapi.staticfiles import StaticFiles
from fastapi.responses import StreamingResponse, JSONResponse
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
from typing import Optional, List, Dict, Any
import sqlite3, json, io, os, math, traceback, secrets, smtplib
from datetime import datetime, date, timedelta
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
import pandas as pd
import numpy as np
from passlib.context import CryptContext
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from fastapi import Depends
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─── App Setup ───────────────────────────────────────────────────────────────
app = FastAPI(title="Picker Planning API", version="1.0")
app.add_middleware(CORSMiddleware, allow_origins=["*"], allow_methods=["*"], allow_headers=["*"])

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DB_PATH  = os.path.join(BASE_DIR, "picker_planning.db")

# ─── Numpy-safe helpers ───────────────────────────────────────────────────────
class _NpEnc(json.JSONEncoder):
    def default(self, o):
        if isinstance(o, np.integer): return int(o)
        if isinstance(o, np.floating): return float(o)
        if isinstance(o, np.bool_):    return bool(o)
        if isinstance(o, np.ndarray):  return o.tolist()
        return super().default(o)

def jdumps(obj): return json.dumps(obj, cls=_NpEnc)
def sanitize(obj):
    if isinstance(obj, dict):  return {k: sanitize(v) for k, v in obj.items()}
    if isinstance(obj, list):  return [sanitize(v) for v in obj]
    if isinstance(obj, np.integer): return int(obj)
    if isinstance(obj, np.floating): return float(obj)
    if isinstance(obj, np.bool_):   return bool(obj)
    return obj

# ─── Database ─────────────────────────────────────────────────────────────────
def get_db():
    conn = sqlite3.connect(DB_PATH, check_same_thread=False)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode=WAL")
    return conn

def init_db():
    conn = get_db()
    conn.executescript("""
    CREATE TABLE IF NOT EXISTS plans (
        token TEXT PRIMARY KEY, plan_date TEXT NOT NULL,
        run_number INTEGER NOT NULL DEFAULT 1, created_at TEXT NOT NULL,
        notes TEXT DEFAULT '', config_json TEXT, demand_json TEXT,
        total_dos INTEGER DEFAULT 0, total_qty INTEGER DEFAULT 0,
        pickers_used INTEGER DEFAULT 0, avg_util REAL DEFAULT 0, skipped_dos INTEGER DEFAULT 0
    );
    CREATE TABLE IF NOT EXISTS plan_details (
        id INTEGER PRIMARY KEY AUTOINCREMENT, token TEXT NOT NULL,
        plan_date TEXT NOT NULL, run_number INTEGER NOT NULL,
        priority INTEGER, do_no TEXT, sto_no TEXT DEFAULT '', st_cd TEXT DEFAULT '',
        st_nm TEXT DEFAULT '', floor INTEGER, sec TEXT DEFAULT '', do_qty INTEGER,
        picker_no INTEGER, machine_no TEXT, scanner_name TEXT DEFAULT '',
        grp TEXT DEFAULT 'G1', bgt_machine INTEGER DEFAULT 3000,
        start_time TEXT, end_time TEXT, duration_min REAL, pcs_per_min REAL,
        cap_used INTEGER, util_pct REAL, remaining INTEGER, over_wh INTEGER DEFAULT 0,
        avail_min REAL DEFAULT 0, status TEXT DEFAULT 'Planned',
        cancel_reason TEXT DEFAULT '', cancelled_at TEXT DEFAULT ''
    );
    CREATE TABLE IF NOT EXISTS picker_day_state (
        plan_date TEXT NOT NULL, machine_no TEXT NOT NULL, floor INTEGER NOT NULL,
        cap_used INTEGER DEFAULT 0, avail_min REAL NOT NULL, last_token TEXT,
        PRIMARY KEY (plan_date, machine_no, floor)
    );
    CREATE TABLE IF NOT EXISTS actual_times (
        id INTEGER PRIMARY KEY AUTOINCREMENT, token TEXT NOT NULL, do_no TEXT NOT NULL,
        plan_date TEXT DEFAULT '', actual_date TEXT DEFAULT '',
        actual_start TEXT DEFAULT '', actual_end TEXT DEFAULT '',
        actual_qty INTEGER DEFAULT 0, notes TEXT DEFAULT '', entered_at TEXT,
        UNIQUE(token, do_no)
    );
    CREATE INDEX IF NOT EXISTS idx_pd_token  ON plan_details(token);
    CREATE INDEX IF NOT EXISTS idx_pd_date   ON plan_details(plan_date);
    CREATE INDEX IF NOT EXISTS idx_act_token ON actual_times(token);
    CREATE INDEX IF NOT EXISTS idx_pds_date  ON picker_day_state(plan_date);
    CREATE TABLE IF NOT EXISTS users (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        username TEXT UNIQUE NOT NULL,
        hashed_pw TEXT NOT NULL,
        role TEXT DEFAULT 'operator',
        email TEXT DEFAULT '',
        created_at TEXT,
        active INTEGER DEFAULT 1
    );
    CREATE TABLE IF NOT EXISTS email_config (
        id INTEGER PRIMARY KEY,
        smtp_host TEXT DEFAULT '',
        smtp_port INTEGER DEFAULT 587,
        smtp_user TEXT DEFAULT '',
        smtp_pass TEXT DEFAULT '',
        from_addr TEXT DEFAULT '',
        enabled INTEGER DEFAULT 0
    );
    CREATE TABLE IF NOT EXISTS email_schedule (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        report_type TEXT NOT NULL,
        frequency TEXT NOT NULL,
        recipients TEXT DEFAULT '',
        enabled INTEGER DEFAULT 1,
        last_sent TEXT DEFAULT ''
    );
    """)
    # Migrate existing DB
    existing = {r[1] for r in conn.execute("PRAGMA table_info(plan_details)").fetchall()}
    for col, defn in [('status',"TEXT DEFAULT 'Planned'"),
                      ('cancel_reason','TEXT DEFAULT ""'),
                      ('cancelled_at','TEXT DEFAULT ""')]:
        if col not in existing:
            conn.execute(f"ALTER TABLE plan_details ADD COLUMN {col} {defn}")
    conn.commit(); conn.close()

init_db()

# ─── Seed default admin if no users exist ────────────────────────────────────
def seed_admin():
    conn = get_db()
    cnt = conn.execute("SELECT COUNT(*) FROM users").fetchone()[0]
    if cnt == 0:
        conn.execute("INSERT INTO users(username,hashed_pw,role,email,created_at) VALUES(?,?,?,?,?)",
                     ('admin', hash_pw('admin123'), 'admin', '', datetime.now().isoformat()))
        conn.commit()
    cnt2 = conn.execute("SELECT COUNT(*) FROM email_config").fetchone()[0]
    if cnt2 == 0:
        conn.execute("INSERT INTO email_config(id,smtp_host,smtp_port,smtp_user,smtp_pass,from_addr,enabled) VALUES(1,'',587,'','','',0)")
        conn.commit()
    conn.close()
seed_admin()

# ─── Status Constants ─────────────────────────────────────────────────────────
STATUS_OPTS = ['Planned', 'Done', 'Delayed', 'Not Picked', 'Cancelled']
STATUS_COLORS = {
    'Done':'#2ECC71','Planned':'#4F8EF7','Delayed':'#F39C12',
    'Not Picked':'#7A8099','Cancelled':'#E74C3C'
}

# ─── Auth ────────────────────────────────────────────────────────────────────
SECRET_KEY = os.environ.get('PICKER_SECRET', secrets.token_hex(32))
pwd_ctx    = CryptContext(schemes=["bcrypt"], deprecated="auto")
bearer     = HTTPBearer(auto_error=False)

ROLES = {'admin': 3, 'operator': 2, 'viewer': 1}

def hash_pw(pw):  return pwd_ctx.hash(pw)
def verify_pw(pw, h): return pwd_ctx.verify(pw, h)

def make_token_jwt(user_id, role):
    import hmac, hashlib, base64
    header  = base64.urlsafe_b64encode(json.dumps({'alg':'HS256','typ':'JWT'}).encode()).rstrip(b'=').decode()
    payload = base64.urlsafe_b64encode(json.dumps({
        'sub': str(user_id), 'role': role,
        'exp': (datetime.utcnow() + timedelta(hours=12)).isoformat()
    }).encode()).rstrip(b'=').decode()
    sig_input = f"{header}.{payload}".encode()
    sig = base64.urlsafe_b64encode(hmac.new(SECRET_KEY.encode(), sig_input, hashlib.sha256).digest()).rstrip(b'=').decode()
    return f"{header}.{payload}.{sig}"

def verify_token_jwt(token):
    try:
        import hmac, hashlib, base64
        parts = token.split('.')
        if len(parts) != 3: return None
        header, payload, sig = parts
        sig_input = f"{header}.{payload}".encode()
        expected = base64.urlsafe_b64encode(hmac.new(SECRET_KEY.encode(), sig_input, hashlib.sha256).digest()).rstrip(b'=').decode()
        if sig != expected: return None
        pad = lambda s: s + '=' * (-len(s) % 4)
        data = json.loads(base64.urlsafe_b64decode(pad(payload)))
        if datetime.fromisoformat(data['exp']) < datetime.utcnow(): return None
        return data
    except: return None

def get_current_user(creds: HTTPAuthorizationCredentials = Depends(bearer)):
    if not creds: return None
    return verify_token_jwt(creds.credentials)

def require_operator(user=Depends(get_current_user)):
    if not user or ROLES.get(user.get('role',''),0) < 2:
        raise HTTPException(401, 'Login required (operator or admin)')
    return user

def require_admin(user=Depends(get_current_user)):
    if not user or ROLES.get(user.get('role',''),0) < 3:
        raise HTTPException(401, 'Admin access required')
    return user

# ─── Time Helpers ─────────────────────────────────────────────────────────────
def m2t(m): t=int(round(m)); return f"{t//60:02d}:{t%60:02d}"
def str2min(s):
    try: h,m=str(s).strip().split(':'); return int(h)*60+int(m)
    except: return None
def dur_min(s,e):
    a,b=str2min(s),str2min(e)
    return (b-a) if (a is not None and b is not None) else None
def adv(clock,dur,ls,le):
    if ls<=clock<le: clock=le
    e=clock+dur
    if clock<ls<e: e+=(le-ls)
    return e
_G_THRESHOLDS = {'g1_min': 3000, 'g2_min': 2000}  # overridden from cfg in allocate()
def auto_group(bgt, g1_min=3000, g2_min=2000):
    bgt=int(bgt)
    if bgt>=g1_min: return 'G1',1
    if bgt>=g2_min: return 'G2',2
    return 'G3',3

# ─── DB Helpers ───────────────────────────────────────────────────────────────
def get_run_number(plan_date):
    conn=get_db()
    row=conn.execute("SELECT COALESCE(MAX(run_number),0) FROM plans WHERE plan_date=?",(plan_date,)).fetchone()
    conn.close(); return int(row[0])+1

def make_token(plan_date,run_number):
    return f"PKP-{plan_date.replace('-','')}-R{run_number:02d}"

def get_allocated_dos(plan_date):
    """DOs actively planned for this date — Cancelled DOs are excluded (free to re-plan)."""
    conn=get_db()
    rows=conn.execute("""
        SELECT DISTINCT do_no FROM plan_details
        WHERE plan_date=? AND COALESCE(status,'Planned') NOT IN ('Cancelled','Deleted')
    """,(plan_date,)).fetchall()
    conn.close(); return {r[0] for r in rows}

def get_globally_locked_dos():
    """ALL DOs that are Planned/Done/Delayed on ANY date. Cancelled = free to re-plan."""
    conn=get_db()
    rows=conn.execute("""
        SELECT DISTINCT do_no FROM plan_details
        WHERE COALESCE(status,'Planned') NOT IN ('Cancelled','Deleted')
    """).fetchall()
    conn.close(); return {r[0] for r in rows}

def save_plan(token,plan_date,run_number,cfg,demand,plan_df,skipped,notes=""):
    conn=get_db()
    avg_u=float(plan_df.groupby('machine_no')['util_pct'].max().mean()) if not plan_df.empty else 0.0
    conn.execute("""
        INSERT OR REPLACE INTO plans
        (token,plan_date,run_number,created_at,notes,config_json,demand_json,
         total_dos,total_qty,pickers_used,avg_util,skipped_dos)
        VALUES(?,?,?,?,?,?,?,?,?,?,?,?)""",(
        token,plan_date,run_number,datetime.now().isoformat(),notes,
        jdumps(sanitize(cfg)),jdumps(sanitize(demand)),
        len(plan_df),
        int(plan_df['do_qty'].sum()) if not plan_df.empty else 0,
        plan_df['machine_no'].nunique() if not plan_df.empty else 0,
        round(avg_u,1),len(skipped)))
    if not plan_df.empty:
        conn.execute("DELETE FROM plan_details WHERE token=?",(token,))
        for _,r in plan_df.iterrows():
            conn.execute("""
                INSERT INTO plan_details
                (token,plan_date,run_number,priority,do_no,sto_no,st_cd,st_nm,
                 floor,sec,do_qty,picker_no,machine_no,scanner_name,grp,bgt_machine,
                 start_time,end_time,duration_min,pcs_per_min,cap_used,util_pct,
                 remaining,over_wh,avail_min)
                VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",(
                token,plan_date,run_number,
                int(r['priority']),str(r['do_no']),str(r.get('sto_no','')),
                str(r.get('st_cd','')),str(r.get('st_nm','')),
                int(r['floor']),str(r['sec']),int(r['do_qty']),
                int(r['picker_no']),str(r['machine_no']),str(r['scanner_name']),
                str(r['grp']),int(r['bgt_machine']),
                str(r['start_time']),str(r['end_time']),
                float(r['duration_min']),float(r['pcs_per_min']),
                int(r['cap_used']),float(r['util_pct']),
                int(r['remaining']),int(r['over_wh']),float(r.get('_avail_min',0))))
    conn.commit(); conn.close()

def save_picker_state(plan_date,plan_df,token):
    if plan_df.empty: return
    conn=get_db()
    state=plan_df.groupby(['machine_no','floor']).agg(
        cap=('cap_used','max'),av=('_avail_min','max')).reset_index()
    for _,r in state.iterrows():
        conn.execute("""
            INSERT INTO picker_day_state(plan_date,machine_no,floor,cap_used,avail_min,last_token)
            VALUES(?,?,?,?,?,?)
            ON CONFLICT(plan_date,machine_no,floor) DO UPDATE SET
              cap_used=MAX(cap_used,excluded.cap_used),
              avail_min=MAX(avail_min,excluded.avail_min),
              last_token=excluded.last_token""",
            (plan_date,str(r['machine_no']),int(r['floor']),int(r['cap']),float(r['av']),token))
    conn.commit(); conn.close()

def load_picker_state(plan_date):
    conn=get_db()
    rows=conn.execute("SELECT machine_no,floor,cap_used,avail_min FROM picker_day_state WHERE plan_date=?",(plan_date,)).fetchall()
    conn.close()
    return {(r['machine_no'],r['floor']):{'cap_used':r['cap_used'],'avail_min':r['avail_min']} for r in rows}

def rebuild_picker_state_for_date(plan_date: str):
    """
    Recompute picker_day_state for a date from scratch using only
    ACTIVE (non-cancelled, non-deleted) DOs in plan_details.

    Called after any DO status change (cancel, delete) so that the next
    allocation run sees the correct avail_min and cap_used — not phantom
    capacity from DOs that no longer exist.

    Logic:
      For each (machine_no, floor) on this date:
        - Collect all plan_details rows that are NOT Cancelled / Deleted
        - max(cap_used) across those rows  → new cap_used
        - max(end_time as minutes) across those rows → new avail_min
        - If zero active DOs remain → delete the state row entirely
          (machine is fully free; next run starts it from shift start)
    """
    conn = get_db()
    # Fetch all active DOs for this date
    rows = conn.execute("""
        SELECT machine_no, floor, cap_used, end_time
        FROM plan_details
        WHERE plan_date=?
          AND COALESCE(status,'Planned') NOT IN ('Cancelled','Deleted')
    """, (plan_date,)).fetchall()

    def t2m(s):
        try:
            h, m = str(s).strip().split(':'); return int(h)*60+int(m)
        except: return 0

    # Group by (machine_no, floor)
    state = {}
    for r in rows:
        key = (str(r['machine_no']), int(r['floor']))
        if key not in state:
            state[key] = {'cap': 0, 'avail': 0}
        state[key]['cap']   = max(state[key]['cap'],   int(r['cap_used'] or 0))
        state[key]['avail'] = max(state[key]['avail'],  t2m(r['end_time']))

    # Get all existing state rows for this date
    existing = conn.execute(
        "SELECT machine_no, floor FROM picker_day_state WHERE plan_date=?",
        (plan_date,)
    ).fetchall()

    # Delete rows that now have zero active DOs (machine fully freed)
    freed = 0
    for er in existing:
        key = (str(er['machine_no']), int(er['floor']))
        if key not in state:
            conn.execute(
                "DELETE FROM picker_day_state WHERE plan_date=? AND machine_no=? AND floor=?",
                (plan_date, er['machine_no'], er['floor'])
            )
            freed += 1

    # Update rows that still have active DOs
    last_tok = conn.execute(
        "SELECT token FROM plans WHERE plan_date=? ORDER BY run_number DESC LIMIT 1",
        (plan_date,)
    ).fetchone()
    tok = last_tok['token'] if last_tok else ''

    for (mno, fl), v in state.items():
        conn.execute("""
            INSERT INTO picker_day_state(plan_date,machine_no,floor,cap_used,avail_min,last_token)
            VALUES(?,?,?,?,?,?)
            ON CONFLICT(plan_date,machine_no,floor) DO UPDATE SET
              cap_used=excluded.cap_used,
              avail_min=excluded.avail_min,
              last_token=excluded.last_token
        """, (plan_date, mno, fl, v['cap'], float(v['avail']), tok))

    conn.commit(); conn.close()
    return {'rebuilt': len(state), 'freed': freed}


def cancel_plan(token):
    """Cancel all DOs in a plan — frees them for re-planning on any date."""

def cancel_plan(token):
    """Cancel all DOs in a plan — frees them for re-planning on any date.
    Also rebuilds picker_day_state so next run gets correct start times."""
    conn=get_db(); now=datetime.now().isoformat()
    # Get plan_date before deleting
    row = conn.execute("SELECT plan_date FROM plans WHERE token=?", (token,)).fetchone()
    plan_date = row['plan_date'] if row else None
    conn.execute("""UPDATE plan_details SET status='Cancelled',
        cancel_reason='Plan cancelled',cancelled_at=? WHERE token=?""",(now,token))
    conn.execute("DELETE FROM plans WHERE token=?",(token,))
    conn.commit(); conn.close()
    # Rebuild state — freed slots are now available for next run
    if plan_date:
        rebuild_picker_state_for_date(plan_date)

def update_do_status(token,do_no,status,cancel_reason=''):
    conn=get_db(); now=datetime.now().isoformat()
    conn.execute("""UPDATE plan_details SET status=?,cancel_reason=?,
        cancelled_at=CASE WHEN ?='Cancelled' THEN ? ELSE cancelled_at END
        WHERE token=? AND do_no=?""",(status,cancel_reason,status,now,token,str(do_no)))
    # Get plan_date for state rebuild
    row = conn.execute("SELECT plan_date FROM plan_details WHERE token=? AND do_no=? LIMIT 1",
                       (token,str(do_no))).fetchone()
    plan_date = row['plan_date'] if row else None
    conn.commit(); conn.close()
    # Rebuild picker_day_state whenever a DO moves to a freeing status
    if plan_date and status in ('Cancelled','Deleted','Not Picked'):
        rebuild_picker_state_for_date(plan_date)

def list_plans(plan_date=None,limit=200):
    conn=get_db()
    if plan_date:
        rows=conn.execute("SELECT * FROM plans WHERE plan_date=? ORDER BY run_number DESC",(plan_date,)).fetchall()
    else:
        rows=conn.execute("SELECT * FROM plans ORDER BY created_at DESC LIMIT ?",(limit,)).fetchall()
    conn.close(); return [dict(r) for r in rows]

def load_plan_details(token):
    conn=get_db()
    rows=conn.execute("SELECT * FROM plan_details WHERE token=? ORDER BY floor,picker_no,start_time",(token,)).fetchall()
    conn.close(); return [dict(r) for r in rows]

def load_actuals_map(token):
    conn=get_db()
    rows=conn.execute("SELECT * FROM actual_times WHERE token=?",(token,)).fetchall()
    conn.close(); return {r['do_no']:dict(r) for r in rows}

def save_actuals(token,records):
    conn=get_db(); now=datetime.now().isoformat()
    for r in records:
        if not r.get('actual_start') or not r.get('actual_end'): continue
        conn.execute("""
            INSERT INTO actual_times
            (token,do_no,plan_date,actual_date,actual_start,actual_end,actual_qty,notes,entered_at)
            VALUES(?,?,?,?,?,?,?,?,?)
            ON CONFLICT(token,do_no) DO UPDATE SET
              actual_date=excluded.actual_date,actual_start=excluded.actual_start,
              actual_end=excluded.actual_end,actual_qty=excluded.actual_qty,
              notes=excluded.notes,entered_at=excluded.entered_at""",(
            token,str(r['do_no']),r.get('plan_date',''),
            r.get('actual_date',str(date.today())),
            str(r['actual_start']),str(r['actual_end']),
            int(r.get('actual_qty',0)),str(r.get('notes','')),now))
    conn.commit(); conn.close()

# ─── Demand Analysis ──────────────────────────────────────────────────────────
def analyse(do_df,cfg):
    FILL=cfg['fill_pct']; BGT=cfg['bgt']
    floors=sorted(do_df['FLOOR'].unique())
    fqty=do_df.groupby('FLOOR')['DO_QTY'].sum()
    fdos=do_df.groupby('FLOOR')['DO_NO'].count()
    rows=[]
    for f in floors:
        qty=int(fqty[f]); dos=int(fdos[f])
        req=max(1,math.ceil(qty/(BGT*FILL/100)))
        exp=round(qty/(req*BGT)*100,1)
        pbd=(do_df[do_df['FLOOR']==f].groupby('PRIORITY')
             .agg(D=('DO_NO','count'),Q=('DO_QTY','sum'))
             .reset_index().sort_values('PRIORITY'))
        rows.append({'floor':int(f),'qty':int(qty),'dos':dos,'req':int(req),'exp':float(exp),
                     'prio':[{'p':int(r['PRIORITY']),'d':int(r['D']),'q':int(r['Q'])} for _,r in pbd.iterrows()]})
    return {'floors':rows,'total_qty':int(do_df['DO_QTY'].sum()),
            'total_dos':len(do_df),'total_req':int(sum(r['req'] for r in rows))}

# ─── Dynamic plan start time ──────────────────────────────────────────────────
def get_plan_start_min(plan_date_str: str, cfg: dict) -> float:
    """
    Future date  → shift start (e.g. 08:00 = 480 min).
    Today        → max(shift_start, current_time_rounded_up_to_5min).
    Past date    → shift start (historical re-plan).
    """
    S = float(cfg['start_min'])
    try:
        plan_dt = date.fromisoformat(plan_date_str)
        today   = date.today()
        if plan_dt == today:
            now = datetime.now()
            now_min = now.hour * 60 + now.minute
            now_min = math.ceil(now_min / 5) * 5   # round UP to next 5-min slot
            return float(max(S, now_min))
    except Exception:
        pass
    return S

# ─── Allocation Engine v2 — Unified mixed pool, no group tier segregation ─────
#
# Key design decisions:
#   1. G1/G2/G3 labels are display-only. No machine is restricted to any
#      priority range. All machines on a floor compete for all DOs.
#
#   2. Machine sort order: BGT descending (highest-capacity first). This means
#      high-capacity machines get lower picker_no labels and are preferred when
#      two machines are equally available.
#
#   3. DO processing order per floor: PRIORITY ASC → SEC ASC → DO_NO ASC.
#      Urgent DOs are always assigned first, guaranteeing priority order.
#
#   4. Picker selection (greedy, earliest-available-first):
#        candidates = machines with (bgt - cap_used) >= DO_QTY
#        sort by: avail_min ASC, then remaining_capacity DESC (tiebreak)
#        assign to candidates[0]
#      "Earliest available" = the machine whose clock is furthest back.
#      This naturally load-balances — a busy machine is skipped until free.
#      The tiebreak favours larger machines when two finish at the same time,
#      keeping smaller machines fresher for future small DOs.
#
#   5. Dynamic start time:
#        - New machine, future plan date   → plan_start = shift_start (08:00)
#        - New machine, today's plan       → plan_start = NOW (rounded up 5 min)
#        - Machine with prior run state    → plan_start = prior run end time
#
def allocate(do_df, machine_df, cfg, demand, prev_state=None, skip_dos=None, plan_date=None):
    S      = cfg['start_min'];  LS = cfg['lunch_s'];  LE = cfg['lunch_e']
    EFF    = cfg['eff_min'];    WH_END = cfg['wh_end'];  BGT_DEF = cfg['bgt']
    if prev_state is None: prev_state = {}
    if skip_dos   is None: skip_dos   = set()

    # Effective start minute for brand-new machines in this run
    plan_start = get_plan_start_min(plan_date or '', cfg)

    # ── Normalise machine DataFrame ───────────────────────────────────────────
    mdf = machine_df.copy()
    mdf.columns = [c.strip().upper().replace(' ','_').replace('/','_') for c in mdf.columns]
    mdf = mdf.loc[:, ~mdf.columns.duplicated(keep='first')]
    sc  = next((c for c in mdf.columns if 'SCAN' in c), None)
    if sc and sc != 'SCANNER_NAME': mdf.rename(columns={sc: 'SCANNER_NAME'}, inplace=True)
    mdf = mdf.drop_duplicates(subset=['MACHINE_NO'], keep='first').reset_index(drop=True)

    bgt_col = next((c for c in ['BGT_PICKER','BGT','BUDGET','CAPACITY'] if c in mdf.columns), None)
    if bgt_col:
        bgt_series = mdf[bgt_col]
        if isinstance(bgt_series, pd.DataFrame): bgt_series = bgt_series.iloc[:, 0]
        mdf['BGT_MACHINE'] = pd.to_numeric(bgt_series, errors='coerce').fillna(BGT_DEF).astype(int)
    else:
        mdf['BGT_MACHINE'] = BGT_DEF

    g1m=int(cfg.get('g1_min',3000)); g2m=int(cfg.get('g2_min',2000))
    mdf['GROUP'], mdf['GROUP_TIER'] = zip(*mdf['BGT_MACHINE'].apply(lambda b: auto_group(b, g1m, g2m)))
    # Sort by BGT descending — highest capacity machines get lowest picker index
    mdf = mdf.sort_values('BGT_MACHINE', ascending=False).reset_index(drop=True)
    all_m = mdf.to_dict('records')

    # ── Distribute machines to floors (demand-weighted) ───────────────────────
    # Each floor gets its required count from the sorted pool.
    # Remaining machines go into an overflow pool shared across all floors.
    pool_map = {}; cur = 0
    for row in sorted(demand['floors'], key=lambda r: -r['req']):
        f   = row['floor']; req = row['req']
        cnt = min(req, len(all_m) - cur)
        pool_map[f] = all_m[cur: cur + cnt]; cur += cnt
    # Overflow: machines not assigned to any floor — available to any floor that runs dry
    overflow_pool = all_m[cur:]

    plans   = []
    skipped = []
    work    = do_df.sort_values(['FLOOR','PRIORITY','SEC','DO_NO']).reset_index(drop=True)

    for floor in sorted(do_df['FLOOR'].unique()):
        pool = pool_map.get(floor, [])
        if not pool:
            for _, dr in work[work['FLOOR'] == floor].iterrows():
                if str(dr['DO_NO']) not in skip_dos:
                    skipped.append((str(dr['DO_NO']), f'No machines assigned to Floor {floor}'))
            continue

        # Build unified picker state — one entry per machine, mix of all groups
        pk = []
        for idx, m in enumerate(pool):
            key  = (str(m['MACHINE_NO']), floor)
            prev = prev_state.get(key, {})
            bgt  = int(m['BGT_MACHINE'])
            # Prior run: continue from where it left off; new machine: use plan_start
            avail = float(prev.get('avail_min', plan_start))
            cap   = int(prev.get('cap_used', 0))
            if (bgt - cap) <= 0 or avail >= WH_END: continue   # skip exhausted machines
            pk.append({'idx': idx, 'cap': cap, 'avail': avail, 'bgt': bgt,
                       'grp': m['GROUP'], 'machine': m})

        if not pk:
            for _, dr in work[work['FLOOR'] == floor].iterrows():
                if str(dr['DO_NO']) not in skip_dos:
                    skipped.append((str(dr['DO_NO']), f'F{floor}: all machines full or shift ended'))
            continue

        # Assign DOs in priority order to the unified pool
        for _, dr in work[work['FLOOR'] == floor].sort_values(['PRIORITY','SEC','DO_NO']).iterrows():
            do_no = str(dr['DO_NO'])
            if do_no in skip_dos: continue
            prio = int(dr['PRIORITY']); qty = int(dr['DO_QTY'])

            # Any machine with enough remaining capacity is a candidate
            cands = [(i, p) for i, p in enumerate(pk) if p['bgt'] - p['cap'] >= qty]
            if not cands:
                # Try pulling from overflow pool — machines not initially assigned to this floor
                for om in overflow_pool:
                    key = (str(om['MACHINE_NO']), floor)
                    if any(p['machine']['MACHINE_NO'] == om['MACHINE_NO'] for p in pk):
                        continue  # already in pool
                    prev = prev_state.get(key, {})
                    bgt  = int(om['BGT_MACHINE'])
                    avail_o = float(prev.get('avail_min', plan_start))
                    cap_o   = int(prev.get('cap_used', 0))
                    if (bgt - cap_o) <= 0 or avail_o >= WH_END: continue
                    new_idx = len(pk)
                    pk.append({'idx': new_idx, 'cap': cap_o, 'avail': avail_o, 'bgt': bgt,
                               'grp': om['GROUP'], 'machine': om})
                cands = [(i, p) for i, p in enumerate(pk) if p['bgt'] - p['cap'] >= qty]
            if not cands:
                skipped.append((do_no, f'F{floor} P{prio}: no machine has {qty} pcs remaining (all {len(pk)} machines full)'))
                continue

            # Earliest available first; tiebreak: largest remaining capacity
            cands.sort(key=lambda x: (x[1]['avail'], -(x[1]['bgt'] - x[1]['cap'])))
            bi, chosen = cands[0]

            ipm = chosen['bgt'] / EFF
            dur = qty / ipm
            ts  = chosen['avail']
            if LS <= ts < LE: ts = LE          # snap to post-lunch if inside window
            te  = adv(ts, dur, LS, LE)

            chosen['cap']   += qty
            chosen['avail']  = te

            plans.append({
                'priority':     prio,
                'do_no':        do_no,
                'sto_no':       str(dr.get('STO_NO', '')),
                'st_cd':        str(dr.get('ST_CD', '')),
                'st_nm':        str(dr.get('ST_NM', '')),
                'floor':        int(floor),
                'sec':          str(dr['SEC']),
                'do_qty':       qty,
                'picker_no':    chosen['idx'] + 1,
                'machine_no':   str(chosen['machine']['MACHINE_NO']),
                'scanner_name': str(chosen['machine'].get('SCANNER_NAME', '')),
                'grp':          chosen['grp'],
                'bgt_machine':  int(chosen['bgt']),
                'start_time':   m2t(ts),
                'end_time':     m2t(te),
                'duration_min': round(dur, 2),
                'pcs_per_min':  round(ipm, 4),
                'cap_used':     int(chosen['cap']),
                'util_pct':     round(chosen['cap'] / chosen['bgt'] * 100, 1),
                'remaining':    int(chosen['bgt'] - chosen['cap']),
                'over_wh':      int(te > WH_END),
                '_avail_min':   float(te),
                'plan_start':   m2t(plan_start),
            })

    return pd.DataFrame(plans), mdf, pool_map, skipped


# ─── Excel Builder ────────────────────────────────────────────────────────────
def make_excel_bytes(plan_df,cfg,meta=None):
    token=(meta or {}).get('token','')
    plan_date=(meta or {}).get('plan_date','')
    run_num=(meta or {}).get('run_number','')
    FILL=cfg.get('fill_pct',70)

    wb=Workbook(); wb.remove(wb.active)
    def hf(c='FFFFFFFF',sz=10): return Font(name='Arial',size=sz,bold=True,color=c)
    def cf(sz=9,b=False,c='FF000000'): return Font(name='Arial',size=sz,bold=b,color=c)
    def fl(h):
        h=h.lstrip('#'); h=('FF'+h) if len(h)==6 else h
        return PatternFill('solid',start_color=h,end_color=h)
    def bd():
        s=Side(style='thin',color='FFBDC3C7')
        return Border(left=s,right=s,top=s,bottom=s)
    def al(h='center'): return Alignment(horizontal=h,vertical='center')

    df=plan_df.copy() if isinstance(plan_df,pd.DataFrame) else pd.DataFrame(plan_df)
    CMAP={'grp':'GROUP','do_qty':'DO_QTY','picker_no':'PICKER_NO','machine_no':'MACHINE_NO',
          'scanner_name':'SCANNER_NAME','bgt_machine':'BGT_MACHINE','start_time':'START_TIME',
          'end_time':'END_TIME','duration_min':'DURATION_MIN','pcs_per_min':'PCS_PER_MIN',
          'cap_used':'CAP_USED','util_pct':'UTIL_%','remaining':'REMAINING','over_wh':'OVER_WH',
          'floor':'FLOOR','sec':'SEC','do_no':'DO_NO','sto_no':'STO_NO','st_cd':'ST_CD',
          'st_nm':'ST_NM','priority':'PRIORITY'}
    df.rename(columns={k:v for k,v in CMAP.items() if k in df.columns},inplace=True)

    # SUMMARY
    ss=wb.create_sheet('SUMMARY',0)
    ss.merge_cells('A1:F1'); ss['A1']=f'PICKER PLAN SUMMARY  |  {token}  |  {plan_date}  |  Run #{run_num}'
    ss['A1'].font=hf(sz=12); ss['A1'].fill=fl('1C2833'); ss['A1'].alignment=al(); ss.row_dimensions[1].height=26
    for i,((lbl,val),bg) in enumerate(zip([
        ('Token',token),('Plan Date',plan_date),('Run #',str(run_num)),
        ('Total DOs',f"{len(df):,}"),
        ('Total QTY',f"{int(df['DO_QTY'].sum()):,}" if 'DO_QTY' in df.columns else '0'),
        ('Fill %',f"{FILL}%")],
        ['1A5276','117A65','6E2FD3','922B21','1F618D','7D6608']),1):
        ss.column_dimensions[get_column_letter(i)].width=18
        lc=ss.cell(row=3,column=i,value=lbl); lc.font=hf(sz=8); lc.fill=fl(bg)
        lc.alignment=al()
        vc=ss.cell(row=4,column=i,value=str(val))
        vc.font=Font(name='Arial',size=13,bold=True,color='FF'+bg); vc.fill=fl('FFFFFF'); vc.alignment=al()
    ss.row_dimensions[3].height=30; ss.row_dimensions[4].height=24; ss.sheet_view.showGridLines=False

    # PICKER_PLAN
    ws=wb.create_sheet('PICKER_PLAN')
    H=['PRIORITY','DO_NO','STO_NO','ST_CD','ST_NM','FLOOR','SEC','DO_QTY',
       'PICKER_NO','MACHINE_NO','SCANNER_NAME','GROUP','BGT_MACHINE',
       'START_TIME','END_TIME','DURATION_MIN','PCS_PER_MIN','CAP_USED','UTIL_%','REMAINING']
    H=[c for c in H if c in df.columns]; NC=len(H)
    ws.merge_cells(f'A1:{get_column_letter(NC)}1')
    ws['A1']=f'PICKER PLAN  |  {token}'; ws['A1'].font=hf(sz=11); ws['A1'].fill=fl('1C2833')
    ws['A1'].alignment=al(); ws.row_dimensions[1].height=24
    for c,(h,w) in enumerate(zip(H,[10,14,14,10,26,7,7,10,10,14,16,8,12,12,12,13,13,12,11,12]),1):
        cell=ws.cell(row=2,column=c,value=h)
        cell.font=hf(); cell.fill=fl('2C3E50'); cell.alignment=al(); cell.border=bd()
        ws.column_dimensions[get_column_letter(c)].width=w
    GCLR={'G1':'FF1E8449','G2':'FF2874A6','G3':'FFF39C12'}
    PC_X={1:('FDEDEC','C0392B'),2:('FEF9E7','F39C12'),3:('EAF4FB','2874A6'),4:('EAFAF1','1E8449')}
    for i,rec in df.iterrows():
        r=3+i; p=int(rec.get('PRIORITY',0)); over=bool(rec.get('OVER_WH',0))
        pbg,_=PC_X.get(p,('F8F9FA','7F8C8D'))
        rf='FFF5E6' if over else ('F0F8FF' if i%2==0 else 'FFFFFF')
        for c,fld in enumerate(H,1):
            cell=ws.cell(row=r,column=c,value=rec.get(fld,''))
            cell.border=bd(); cell.fill=fl(rf); cell.font=cf(); cell.alignment=al()
        ws.cell(row=r,column=1).fill=fl(pbg)
        if 'GROUP' in H:
            gi=H.index('GROUP')+1
            ws.cell(row=r,column=gi).font=Font(name='Arial',size=9,bold=True,color=GCLR.get(str(rec.get('GROUP','G1')),'FF000000'))
        ws.row_dimensions[r].height=15
    ws.freeze_panes='A3'; ws.sheet_view.showGridLines=False

    # PICKER_SCHEDULE
    ps=wb.create_sheet('PICKER_SCHEDULE')
    if not df.empty and 'PICKER_NO' in df.columns:
        sched=df.groupby(['FLOOR','PICKER_NO','MACHINE_NO']).agg(
            DOs=('DO_NO','count'),QTY=('DO_QTY','sum'),
            START=('START_TIME','min'),END=('END_TIME','max'),
            CAP=('CAP_USED','max'),UTIL=('UTIL_%','max')).reset_index()
        SH=['FLOOR','PICKER_NO','MACHINE_NO','DOs','QTY','START','END','CAP','UTIL']
        ps.merge_cells(f'A1:{get_column_letter(len(SH))}1')
        ps['A1']=f'PICKER SCHEDULE  |  {token}'; ps['A1'].font=hf(sz=11); ps['A1'].fill=fl('1C2833'); ps['A1'].alignment=al()
        for c,h in enumerate(SH,1):
            cell=ps.cell(row=2,column=c,value=h)
            cell.font=hf(); cell.fill=fl('2C3E50'); cell.alignment=al(); cell.border=bd()
            ps.column_dimensions[get_column_letter(c)].width=14
        for i,rec in sched.iterrows():
            r=3+i
            for c,fld in enumerate(SH,1):
                cell=ps.cell(row=r,column=c,value=rec.get(fld,''))
                cell.border=bd(); cell.font=cf(); cell.alignment=al()
            ps.row_dimensions[r].height=15
        ps.freeze_panes='A3'; ps.sheet_view.showGridLines=False

    # ACTUAL_TIME_ENTRY template
    at=wb.create_sheet('ACTUAL_TIME_ENTRY')
    ACH=['DO_NO','FLOOR','SEC','PRIORITY','DO_QTY','PICKER_NO','MACHINE_NO',
         'PLAN_START','PLAN_END','ACTUAL_DATE','ACTUAL_START','ACTUAL_END','NOTES']
    at.merge_cells(f'A1:{get_column_letter(len(ACH))}1')
    at['A1']=f'ACTUAL TIME ENTRY  |  {token}  — Fill ACTUAL_DATE/START/END columns'
    at['A1'].font=hf(sz=11); at['A1'].fill=fl('1C2833'); at['A1'].alignment=al(); at.row_dimensions[1].height=24
    for c,h in enumerate(ACH,1):
        cell=at.cell(row=2,column=c,value=h)
        cell.font=hf(c='FFFFD700' if h.startswith('ACTUAL') else 'FFFFFFFF'); cell.fill=fl('2C3E50')
        cell.alignment=al(); cell.border=bd()
        at.column_dimensions[get_column_letter(c)].width=15
    if not df.empty:
        plan_cols=['DO_NO','FLOOR','SEC','PRIORITY','DO_QTY','PICKER_NO','MACHINE_NO','START_TIME','END_TIME']
        plan_cols=[c for c in plan_cols if c in df.columns]
        for i,rec in df.sort_values(['FLOOR','PICKER_NO','START_TIME'] if 'START_TIME' in df.columns else ['FLOOR']).iterrows():
            r=3+i
            vals=[rec.get('DO_NO',''),rec.get('FLOOR',''),rec.get('SEC',''),
                  rec.get('PRIORITY',''),rec.get('DO_QTY',''),rec.get('PICKER_NO',''),
                  rec.get('MACHINE_NO',''),rec.get('START_TIME',''),rec.get('END_TIME',''),'','','','']
            for c,v in enumerate(vals,1):
                cell=at.cell(row=r,column=c,value=v)
                cell.border=bd(); cell.font=cf()
                cell.fill=fl('FFFFF0' if c>=10 else ('F0F8FF' if i%2==0 else 'FFFFFF'))
                cell.alignment=al()
            at.row_dimensions[r].height=15
    at.freeze_panes='A3'; at.sheet_view.showGridLines=False

    buf=io.BytesIO(); wb.save(buf); buf.seek(0); return buf.getvalue()

def make_actuals_template_bytes(token):
    details=load_plan_details(token)
    if not details: return None
    df=pd.DataFrame(details)
    plan_meta_rows=list_plans()
    meta=next((p for p in plan_meta_rows if p['token']==token),{})
    cfg=json.loads(meta.get('config_json') or '{"fill_pct":70,"bgt":3000}')
    return make_excel_bytes(df,cfg,meta)

def make_status_template_bytes(token):
    details=load_plan_details(token)
    if not details: return None
    wb=Workbook(); ws=wb.active; ws.title='STATUS_UPDATE'
    def fl(h): h=h.lstrip('#'); return PatternFill('solid',start_color=('FF'+h),end_color=('FF'+h))
    def hf(): return Font(name='Arial',size=9,bold=True,color='FF4F8EF7')
    cols=['DO_NO','FLOOR','SEC','PRIORITY','DO_QTY','PICKER_NO','PLAN_START','PLAN_END','CURRENT_STATUS','NEW_STATUS','CANCEL_REASON']
    for ci,col in enumerate(cols,1):
        c=ws.cell(row=1,column=ci,value=col); c.font=hf(); c.fill=fl('1C2836')
        c.alignment=Alignment(horizontal='center'); ws.column_dimensions[get_column_letter(ci)].width=16
    ws.cell(row=2,column=10,value='Options: Done|Cancelled|Not Picked|Delayed|Planned').fill=fl('0D1520')
    ws.cell(row=2,column=11,value='Picker absent|Stock unavailable|DO cancelled by WH|End of shift|Other').fill=fl('0D1520')
    for rec in sorted(details,key=lambda r:(r.get('floor',0),r.get('picker_no',0),r.get('start_time',''))):
        ws.append([str(rec.get('do_no','')),rec.get('floor',0),rec.get('sec',''),
                   rec.get('priority',1),rec.get('do_qty',0),rec.get('picker_no',0),
                   rec.get('start_time',''),rec.get('end_time',''),
                   rec.get('status','Planned'),'',''])
    buf=io.BytesIO(); wb.save(buf); buf.seek(0); return buf.getvalue()

# ─── Pydantic Models ──────────────────────────────────────────────────────────
class GenerateRequest(BaseModel):
    plan_date: str
    notes: str = ""
    cfg: Dict[str,Any]
    dos: List[Dict[str,Any]]
    machines: List[Dict[str,Any]]

class ActualRecord(BaseModel):
    do_no: str
    plan_date: str = ""
    actual_date: str = ""
    actual_start: str
    actual_end: str
    actual_qty: int = 0
    notes: str = ""

class StatusUpdate(BaseModel):
    do_no: str
    status: str
    cancel_reason: str = ""

# ─── API Routes ───────────────────────────────────────────────────────────────

@app.post("/api/excel/parse")
async def parse_excel(file: UploadFile = File(...), fill_pct: float = 70, bgt: int = 3000):
    """Parse DO+Machine Excel, return JSON."""
    try:
        content = await file.read()
        xl = pd.ExcelFile(io.BytesIO(content))
        sheets = xl.sheet_names

        do_sheet = next((s for s in sheets if 'DO' in s.upper() and 'MACHINE' not in s.upper()), sheets[0])
        mac_sheet = next((s for s in sheets if 'MACHINE' in s.upper() or 'MAC' in s.upper()), sheets[min(1,len(sheets)-1)])

        do_df = pd.read_excel(io.BytesIO(content), sheet_name=do_sheet)
        mc_df = pd.read_excel(io.BytesIO(content), sheet_name=mac_sheet)

        do_df.columns  = [str(c).strip().upper().replace(' ','_') for c in do_df.columns]
        mc_df.columns  = [str(c).strip().upper().replace(' ','_') for c in mc_df.columns]
        do_df = do_df.dropna(subset=['DO_NO'])

        # Normalise required cols
        for c in ['PRIORITY','DO_QTY','FLOOR']:
            if c in do_df.columns:
                do_df[c] = pd.to_numeric(do_df[c], errors='coerce').fillna(0).astype(int)
        if 'SEC' not in do_df.columns and 'SECTION' in do_df.columns:
            do_df.rename(columns={'SECTION':'SEC'}, inplace=True)
        if 'SEC' not in do_df.columns:
            do_df['SEC'] = 'A'

        mc_df.columns = [str(c).strip().upper().replace(' ','_').replace('/','_') for c in mc_df.columns]
        # Deduplicate columns
        mc_df = mc_df.loc[:, ~mc_df.columns.duplicated(keep='first')]
        mc_df = mc_df.dropna(subset=['MACHINE_NO'])
        mc_df = mc_df.drop_duplicates(subset=['MACHINE_NO'], keep='first').reset_index(drop=True)

        # Find BGT column
        bgt_col = next((c for c in ['BGT_PICKER','BGT','BUDGET','CAPACITY'] if c in mc_df.columns), None)
        BGT_DEF_PARSE = 3000
        if bgt_col:
            bgt_s = mc_df[bgt_col]
            if isinstance(bgt_s, pd.DataFrame): bgt_s = bgt_s.iloc[:,0]
            mc_df['BGT_PICKER'] = pd.to_numeric(bgt_s, errors='coerce').fillna(BGT_DEF_PARSE).astype(int)
        else:
            mc_df['BGT_PICKER'] = BGT_DEF_PARSE

        # Compute groups
        mc_df['GROUP'] = mc_df['BGT_PICKER'].apply(lambda b: 'G1' if b>=3000 else ('G2' if b>=2000 else 'G3'))

        # Machine availability summary per floor (if FLOOR column exists in machine sheet)
        floor_avail = {}
        if 'FLOOR' in mc_df.columns:
            for f, grp in mc_df.groupby('FLOOR'):
                floor_avail[int(f)] = {'machines': len(grp), 'G1': int((grp['GROUP']=='G1').sum()),
                    'G2': int((grp['GROUP']=='G2').sum()), 'G3': int((grp['GROUP']=='G3').sum()),
                    'avg_bgt': int(grp['BGT_PICKER'].mean())}

        # Group summary
        grp_summary = {}
        for g, grp in mc_df.groupby('GROUP'):
            grp_summary[g] = {'machines': len(grp), 'avg_bgt': int(grp['BGT_PICKER'].mean()),
                              'bgt_sum': int(grp['BGT_PICKER'].sum())}

        # Required machines estimate per floor — use caller's fill_pct and bgt
        floor_required = {}
        if 'FLOOR' in do_df.columns:
            bgt_def = bgt if bgt > 0 else (int(mc_df['BGT_PICKER'].mean()) if len(mc_df) else 3000)
            for f, grp in do_df.groupby('FLOOR'):
                qty = int(grp['DO_QTY'].sum())
                req = max(1, math.ceil(qty / (bgt_def * fill_pct / 100)))
                avail = floor_avail.get(int(f), {}).get('machines', len(mc_df))
                floor_required[int(f)] = {'required': req, 'available': avail, 'surplus': avail - req}

        summary = {
            'total_dos': len(do_df), 'total_qty': int(do_df['DO_QTY'].sum()),
            'floors': sorted([int(f) for f in do_df['FLOOR'].unique()]),
            'priorities': sorted([int(p) for p in do_df['PRIORITY'].unique()]),
            'machines': len(mc_df), 'total_machines': len(mc_df),
            'do_sheet': do_sheet, 'mac_sheet': mac_sheet,
            'g1_count': int((mc_df['GROUP']=='G1').sum()),
            'g2_count': int((mc_df['GROUP']=='G2').sum()),
            'g3_count': int((mc_df['GROUP']=='G3').sum()),
            'avg_bgt': int(mc_df['BGT_PICKER'].mean()),
        }
        return {
            'dos': json.loads(do_df.to_json(orient='records')),
            'machines': json.loads(mc_df.to_json(orient='records')),
            'summary': summary,
            'grp_summary': grp_summary,
            'floor_avail': floor_avail,
            'floor_required': floor_required,
        }
    except Exception as e:
        raise HTTPException(400, f"Excel parse error: {e}")

@app.post("/api/plans/generate")
async def generate_plan(req: GenerateRequest):
    """Run allocation, save to DB, return full plan."""
    try:
        cfg = req.cfg
        # Build default config from provided values
        cfg.setdefault('start_min', 480)
        cfg.setdefault('wh_end', 1020)
        cfg.setdefault('lunch_s', 780)
        cfg.setdefault('lunch_e', 825)
        cfg.setdefault('eff_min', 495)
        cfg.setdefault('bgt', 3000)
        cfg.setdefault('fill_pct', 70)

        do_df  = pd.DataFrame(req.dos)
        mac_df = pd.DataFrame(req.machines)

        # Normalise columns
        do_df.columns  = [c.strip().upper().replace(' ','_') for c in do_df.columns]
        mac_df.columns = [c.strip().upper().replace(' ','_') for c in mac_df.columns]
        for c in ['PRIORITY','DO_QTY','FLOOR']:
            if c in do_df.columns:
                do_df[c] = pd.to_numeric(do_df[c], errors='coerce').fillna(0).astype(int)

        plan_date = req.plan_date
        run_number = get_run_number(plan_date)

        # Global lock: exclude DOs that are Planned/Done/Delayed anywhere
        locked_dos    = get_globally_locked_dos()
        already_today = get_allocated_dos(plan_date)  # already this date (active)

        # Further exclude DOs from other dates
        all_excluded  = locked_dos  # already includes today's active ones
        do_rem = do_df[~do_df['DO_NO'].astype(str).isin(all_excluded)].copy()

        if do_rem.empty:
            locked_count  = len(locked_dos & set(do_df['DO_NO'].astype(str)))
            today_count   = len(already_today)
            return {
                'error': True,
                'message': (f'All {len(do_df)} DOs are already planned/done. '
                            f'{today_count} active today, {locked_count - today_count} on other dates. '
                            f'Cancel existing plans first to re-plan.'),
                'locked_count': locked_count,
                'today_count': today_count
            }

        demand = analyse(do_rem, cfg)
        prev_state = load_picker_state(plan_date)
        plan_df, mdf, pool_map, skipped = allocate(do_rem, mac_df, cfg, demand, prev_state, skip_dos=all_excluded, plan_date=plan_date)

        if plan_df.empty:
            return {'error': True, 'message': f'Allocation produced no results. Skipped: {len(skipped)} DOs.', 'skipped': skipped}

        token = make_token(plan_date, run_number)
        save_plan(token, plan_date, run_number, cfg, demand, plan_df, skipped, req.notes)
        save_picker_state(plan_date, plan_df, token)

        details = sanitize(plan_df.drop(columns=['_avail_min'],errors='ignore').to_dict('records'))

        # Group/machine summary
        grp_summary = {}
        for _, r in mdf.iterrows():
            g = str(r['GROUP']); g_prev = grp_summary.get(g, {'machines':0,'avg_bgt':0,'bgt_sum':0})
            g_prev['machines'] += 1; g_prev['bgt_sum'] += int(r['BGT_MACHINE'])
            grp_summary[g] = g_prev
        for g in grp_summary:
            m = grp_summary[g]['machines']
            grp_summary[g]['avg_bgt'] = grp_summary[g]['bgt_sum'] // m if m else 0

        return {
            'error': False,
            'token': token,
            'plan_date': plan_date,
            'run_number': run_number,
            'total_dos': len(plan_df),
            'total_qty': int(plan_df['do_qty'].sum()),
            'pickers_used': plan_df['machine_no'].nunique(),
            'avg_util': round(float(plan_df.groupby('machine_no')['util_pct'].max().mean()), 1),
            'skipped': len(skipped),
            'skipped_list': [{'do_no': s[0], 'reason': s[1]} for s in skipped],
            'excluded_count': len(do_df) - len(do_rem),
            'plan_start': plan_df['plan_start'].iloc[0] if not plan_df.empty and 'plan_start' in plan_df.columns else cfg.get('start_str','08:00'),
            'demand': sanitize(demand),
            'grp_summary': grp_summary,
            'details': details,
        }
    except Exception as e:
        traceback.print_exc()
        raise HTTPException(500, f"Generation error: {e}")

@app.get("/api/plans")
def api_list_plans(date: Optional[str] = None):
    plans = list_plans(date)
    # Enrich each plan with status breakdown
    conn = get_db()
    for p in plans:
        rows = conn.execute("""
            SELECT COALESCE(status,'Planned') as s, COUNT(*) as c
            FROM plan_details WHERE token=? GROUP BY s""", (p['token'],)).fetchall()
        p['status_breakdown'] = {r['s']: r['c'] for r in rows}
    conn.close()
    return plans

@app.get("/api/plans/{token}")
def api_get_plan(token: str):
    plans = list_plans()
    meta  = next((p for p in plans if p['token'] == token), None)
    if not meta: raise HTTPException(404, "Plan not found")
    details = load_plan_details(token)
    actuals = load_actuals_map(token)
    for d in details:
        act = actuals.get(str(d['do_no']),{})
        d['actual_start'] = act.get('actual_start','')
        d['actual_end']   = act.get('actual_end','')
        d['actual_date']  = act.get('actual_date','')
        d['actual_notes'] = act.get('notes','')
    return {'meta': meta, 'details': details}

@app.delete("/api/plans/{token}")
def api_cancel_plan(token: str):
    """Cancel plan → all DOs become free for re-planning."""
    cancel_plan(token)
    return {'success': True, 'message': f'Plan {token} cancelled. All DOs are now free to re-plan.'}

@app.post("/api/plans/{token}/rebuild-state")
def api_rebuild_state(token: str):
    """Manually rebuild picker_day_state for this plan's date.
    Use this if state seems wrong after bulk cancel/status operations."""
    plans = list_plans()
    meta  = next((p for p in plans if p['token'] == token), None)
    if not meta: raise HTTPException(404, "Plan not found")
    result = rebuild_picker_state_for_date(meta['plan_date'])
    return {'success': True, 'plan_date': meta['plan_date'], **result}

@app.get("/api/plans/{token}/excel")
def api_download_excel(token: str):
    details = load_plan_details(token)
    plans   = list_plans()
    meta    = next((p for p in plans if p['token'] == token), {})
    cfg     = json.loads(meta.get('config_json') or '{"fill_pct":70,"bgt":3000}')
    df      = pd.DataFrame(details)
    xlsx    = make_excel_bytes(df, cfg, meta)
    return StreamingResponse(io.BytesIO(xlsx),
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename="Picker_{token}.xlsx"'})

@app.get("/api/plans/{token}/actuals-template")
def api_actuals_template(token: str):
    xlsx = make_actuals_template_bytes(token)
    if not xlsx: raise HTTPException(404, "No details for token")
    return StreamingResponse(io.BytesIO(xlsx),
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename="Actuals_Template_{token}.xlsx"'})

@app.get("/api/plans/{token}/status-template")
def api_status_template(token: str):
    xlsx = make_status_template_bytes(token)
    if not xlsx: raise HTTPException(404, "No details for token")
    return StreamingResponse(io.BytesIO(xlsx),
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename="Status_Template_{token}.xlsx"'})

@app.post("/api/actuals/{token}")
def api_save_actuals(token: str, records: List[ActualRecord]):
    """Save actual times + auto-update status to Done/Delayed."""
    plans = list_plans()
    meta  = next((p for p in plans if p['token'] == token), None)
    if not meta: raise HTTPException(404, "Plan not found")
    plan_date = meta.get('plan_date','')
    recs = [r.dict() for r in records]
    save_actuals(token, recs)
    # Auto-set status
    for r in recs:
        if r.get('actual_start') and r.get('actual_end'):
            ad = r.get('actual_date', '')
            st = 'Delayed' if (ad and plan_date and ad > plan_date) else 'Done'
            update_do_status(token, r['do_no'], st)
    return {'success': True, 'saved': len([r for r in recs if r.get('actual_start')])}

@app.post("/api/actuals/{token}/bulk-upload")
async def api_bulk_actuals(token: str, file: UploadFile = File(...)):
    """Parse filled actuals Excel template and save."""
    content = await file.read()
    try:
        xl = pd.read_excel(io.BytesIO(content), sheet_name='ACTUAL_TIME_ENTRY')
        xl.columns = [c.strip().upper() for c in xl.columns]
        xl = xl[xl['DO_NO'].notna() & (xl['DO_NO'].astype(str).str.strip()!='')]
        filled = xl[xl['ACTUAL_START'].notna() & xl['ACTUAL_END'].notna() &
                   (xl['ACTUAL_START'].astype(str).str.strip()!='') &
                   (xl['ACTUAL_END'].astype(str).str.strip()!='')]
        if filled.empty:
            return {'success': False, 'message': 'No rows with ACTUAL_START and ACTUAL_END filled.'}
        plans = list_plans()
        meta  = next((p for p in plans if p['token'] == token), None)
        plan_date = meta.get('plan_date','') if meta else ''
        recs = []
        for _,row in filled.iterrows():
            recs.append({
                'do_no': str(row['DO_NO']), 'plan_date': plan_date,
                'actual_date': str(row.get('ACTUAL_DATE','') or '').split(' ')[0] or str(date.today()),
                'actual_start': str(row['ACTUAL_START']).strip(),
                'actual_end':   str(row['ACTUAL_END']).strip(),
                'actual_qty': int(row.get('DO_QTY') or row.get('ACTUAL_QTY') or 0),
                'notes': str(row.get('NOTES','') or '')
            })
        save_actuals(token, recs)
        for r in recs:
            ad = r['actual_date']; st = 'Delayed' if (ad and plan_date and ad > plan_date) else 'Done'
            update_do_status(token, r['do_no'], st)
        return {'success': True, 'saved': len(recs)}
    except Exception as e:
        raise HTTPException(400, f"Upload error: {e}")

@app.post("/api/status/{token}")
def api_update_status(token: str, updates: List[StatusUpdate]):
    """Bulk status update. Rebuilds picker_day_state once at the end (not per-row)."""
    conn = get_db(); now = datetime.now().isoformat()
    count = 0; plan_date = None; needs_rebuild = False
    for u in updates:
        if u.status not in STATUS_OPTS: continue
        conn.execute("""UPDATE plan_details SET status=?,cancel_reason=?,
            cancelled_at=CASE WHEN ?='Cancelled' THEN ? ELSE cancelled_at END
            WHERE token=? AND do_no=?""",
            (u.status, u.cancel_reason, u.status, now, token, str(u.do_no)))
        count += 1
        if u.status in ('Cancelled','Deleted','Not Picked'): needs_rebuild = True
    if not plan_date:
        row = conn.execute("SELECT plan_date FROM plans WHERE token=? LIMIT 1",(token,)).fetchone()
        if row: plan_date = row['plan_date']
    conn.commit(); conn.close()
    # Single rebuild after all updates
    if needs_rebuild and plan_date:
        rebuild_picker_state_for_date(plan_date)
    return {'success': True, 'updated': count}

@app.post("/api/status/{token}/bulk-upload")
async def api_bulk_status(token: str, file: UploadFile = File(...)):
    content = await file.read()
    try:
        xl = pd.read_excel(io.BytesIO(content), sheet_name='STATUS_UPDATE')
        xl.columns = [c.strip().upper() for c in xl.columns]
        xl = xl[xl['DO_NO'].notna() & (xl['DO_NO'].astype(str).str.strip()!='')]
        filled = xl[xl['NEW_STATUS'].notna() & (xl['NEW_STATUS'].astype(str).str.strip()!='')]
        if filled.empty:
            return {'success': False, 'message': 'No rows with NEW_STATUS filled.'}
        conn2 = get_db(); now2 = datetime.now().isoformat()
        count = 0; needs_rebuild = False; plan_date2 = None
        for _,row in filled.iterrows():
            ns = str(row['NEW_STATUS']).strip()
            cr = str(row.get('CANCEL_REASON','') or '').strip()
            if cr in ('nan','None'): cr = ''
            if ns not in STATUS_OPTS: continue
            conn2.execute("""UPDATE plan_details SET status=?,cancel_reason=?,
                cancelled_at=CASE WHEN ?='Cancelled' THEN ? ELSE cancelled_at END
                WHERE token=? AND do_no=?""",
                (ns, cr, ns, now2, token, str(row['DO_NO'])))
            count += 1
            if ns in ('Cancelled','Deleted','Not Picked'): needs_rebuild = True
        if not plan_date2:
            r2 = conn2.execute("SELECT plan_date FROM plans WHERE token=? LIMIT 1",(token,)).fetchone()
            if r2: plan_date2 = r2['plan_date']
        conn2.commit(); conn2.close()
        if needs_rebuild and plan_date2:
            rebuild_picker_state_for_date(plan_date2)
        return {'success': True, 'updated': count}
    except Exception as e:
        raise HTTPException(400, f"Upload error: {e}")

@app.get("/api/analytics/{token}")
def api_analytics(token: str):
    plans = list_plans()
    meta  = next((p for p in plans if p['token'] == token), None)
    if not meta: raise HTTPException(404, "Plan not found")
    details  = load_plan_details(token)
    actuals  = load_actuals_map(token)
    plan_date = meta.get('plan_date','')

    df = pd.DataFrame(details)
    if df.empty: return {'error': 'No details'}

    df['do_no'] = df['do_no'].astype(str)
    df['actual_start'] = df['do_no'].map(lambda x: actuals.get(x,{}).get('actual_start',''))
    df['actual_end']   = df['do_no'].map(lambda x: actuals.get(x,{}).get('actual_end',''))
    df['actual_date']  = df['do_no'].map(lambda x: actuals.get(x,{}).get('actual_date',''))

    df['status'] = df['status'].fillna('Planned')

    status_counts = df['status'].value_counts().to_dict()
    total = len(df)
    done  = int(status_counts.get('Done',0))
    cancelled = int(status_counts.get('Cancelled',0))
    completion_pct = round(done/max(total-cancelled,1)*100,1)

    # Floor breakdown
    floor_data = []
    for floor, fdf in df.groupby('floor'):
        sc = fdf['status'].value_counts().to_dict()
        floor_data.append({
            'floor': int(floor), 'total': len(fdf),
            'qty': int(fdf['do_qty'].sum()),
            'done': sc.get('Done',0), 'delayed': sc.get('Delayed',0),
            'cancelled': sc.get('Cancelled',0), 'planned': sc.get('Planned',0),
            'not_picked': sc.get('Not Picked',0)
        })

    # Picker performance
    picker_data = []
    for (floor,picker), pdf in df.groupby(['floor','picker_no']):
        sc = pdf['status'].value_counts().to_dict()
        machine = pdf['machine_no'].iloc[0] if len(pdf) else ''
        picker_data.append({
            'floor': int(floor), 'picker': int(picker), 'machine': str(machine),
            'dos': len(pdf), 'qty': int(pdf['do_qty'].sum()),
            'util': round(float(pdf['util_pct'].max()),1),
            'done': sc.get('Done',0), 'cancelled': sc.get('Cancelled',0),
            'delayed': sc.get('Delayed',0)
        })

    return {
        'meta': meta, 'status_counts': status_counts,
        'completion_pct': completion_pct, 'total': total,
        'floor_data': floor_data, 'picker_data': picker_data,
        'status_colors': STATUS_COLORS
    }

@app.get("/api/locked-dos")
def api_locked_dos():
    locked = get_globally_locked_dos()
    return {'count': len(locked), 'dos': list(locked)[:100]}

# ─── Auth Endpoints ───────────────────────────────────────────────────────────
@app.post("/api/auth/login")
def api_login(body: dict):
    username = str(body.get('username',''))
    password = str(body.get('password',''))
    conn = get_db()
    row = conn.execute("SELECT * FROM users WHERE username=? AND active=1", (username,)).fetchone()
    conn.close()
    if not row or not verify_pw(password, row['hashed_pw']):
        raise HTTPException(401, "Invalid username or password")
    token = make_token_jwt(row['id'], row['role'])
    return {'token': token, 'username': row['username'], 'role': row['role']}

@app.get("/api/auth/me")
def api_me(user=Depends(get_current_user)):
    if not user: raise HTTPException(401, "Not authenticated")
    return user

# ─── User Management ──────────────────────────────────────────────────────────
@app.get("/api/users")
def api_list_users(user=Depends(require_admin)):
    conn = get_db()
    rows = conn.execute("SELECT id,username,role,email,created_at,active FROM users ORDER BY id").fetchall()
    conn.close()
    return [dict(r) for r in rows]

@app.post("/api/users")
def api_create_user(body: dict, user=Depends(require_admin)):
    username = str(body.get('username','')).strip()
    password = str(body.get('password','admin123')).strip()
    role     = str(body.get('role','operator'))
    email    = str(body.get('email','')).strip()
    if not username: raise HTTPException(400, "Username required")
    if role not in ROLES: raise HTTPException(400, f"Role must be one of: {list(ROLES)}")
    conn = get_db()
    try:
        conn.execute("INSERT INTO users(username,hashed_pw,role,email,created_at) VALUES(?,?,?,?,?)",
                     (username, hash_pw(password), role, email, datetime.now().isoformat()))
        conn.commit()
    except Exception as e:
        conn.close(); raise HTTPException(400, f"Username already exists: {e}")
    conn.close()
    return {'success': True, 'username': username, 'role': role}

@app.put("/api/users/{uid}")
def api_update_user(uid: int, body: dict, user=Depends(require_admin)):
    conn = get_db()
    if 'password' in body and body['password']:
        conn.execute("UPDATE users SET hashed_pw=? WHERE id=?", (hash_pw(body['password']), uid))
    if 'role' in body:
        conn.execute("UPDATE users SET role=? WHERE id=?", (body['role'], uid))
    if 'email' in body:
        conn.execute("UPDATE users SET email=? WHERE id=?", (body['email'], uid))
    if 'active' in body:
        conn.execute("UPDATE users SET active=? WHERE id=?", (int(body['active']), uid))
    conn.commit(); conn.close()
    return {'success': True}

@app.delete("/api/users/{uid}")
def api_delete_user(uid: int, user=Depends(require_admin)):
    conn = get_db()
    conn.execute("UPDATE users SET active=0 WHERE id=?", (uid,))
    conn.commit(); conn.close()
    return {'success': True}

# ─── Email Config ──────────────────────────────────────────────────────────────
@app.get("/api/email/config")
def api_get_email_config(user=Depends(require_admin)):
    conn = get_db()
    row = conn.execute("SELECT * FROM email_config WHERE id=1").fetchone()
    conn.close()
    if not row: return {}
    d = dict(row); d.pop('smtp_pass', None)  # never return password
    return d

@app.post("/api/email/config")
def api_save_email_config(body: dict, user=Depends(require_admin)):
    conn = get_db()
    conn.execute("""UPDATE email_config SET smtp_host=?,smtp_port=?,smtp_user=?,
        smtp_pass=CASE WHEN ?!='' THEN ? ELSE smtp_pass END,
        from_addr=?,enabled=? WHERE id=1""",
        (body.get('smtp_host',''), int(body.get('smtp_port',587)),
         body.get('smtp_user',''), body.get('smtp_pass',''), body.get('smtp_pass',''),
         body.get('from_addr',''), int(body.get('enabled',0))))
    conn.commit(); conn.close()
    return {'success': True}

@app.post("/api/email/test")
def api_test_email(body: dict, user=Depends(require_admin)):
    conn = get_db()
    cfg_row = conn.execute("SELECT * FROM email_config WHERE id=1").fetchone()
    conn.close()
    if not cfg_row or not cfg_row['smtp_host']:
        raise HTTPException(400, "Email not configured")
    try:
        send_email(dict(cfg_row), body.get('to',''), 'PickerOS Test Email', '<p>Test from PickerOS</p>')
        return {'success': True}
    except Exception as e:
        raise HTTPException(500, str(e))

@app.get("/api/email/schedules")
def api_get_schedules(user=Depends(require_admin)):
    conn = get_db()
    rows = conn.execute("SELECT * FROM email_schedule ORDER BY id").fetchall()
    conn.close()
    return [dict(r) for r in rows]

@app.post("/api/email/schedules")
def api_save_schedule(body: dict, user=Depends(require_admin)):
    conn = get_db()
    if body.get('id'):
        conn.execute("UPDATE email_schedule SET report_type=?,frequency=?,recipients=?,enabled=? WHERE id=?",
                     (body['report_type'],body['frequency'],body.get('recipients',''),int(body.get('enabled',1)),body['id']))
    else:
        conn.execute("INSERT INTO email_schedule(report_type,frequency,recipients,enabled) VALUES(?,?,?,?)",
                     (body['report_type'],body['frequency'],body.get('recipients',''),int(body.get('enabled',1))))
    conn.commit(); conn.close()
    return {'success': True}

@app.delete("/api/email/schedules/{sid}")
def api_del_schedule(sid: int, user=Depends(require_admin)):
    conn = get_db(); conn.execute("DELETE FROM email_schedule WHERE id=?", (sid,)); conn.commit(); conn.close()
    return {'success': True}

# ─── Dashboard Endpoint ────────────────────────────────────────────────────────
@app.get("/api/dashboard")
def api_dashboard():
    conn = get_db()
    today = date.today().isoformat()
    yesterday = (date.today() - timedelta(days=1)).isoformat()
    week_start = (date.today() - timedelta(days=date.today().weekday())).isoformat()
    month_start = date.today().replace(day=1).isoformat()

    def safe(q, *a):
        try: return conn.execute(q,a).fetchone()[0] or 0
        except: return 0

    result = {
        'today': {
            'plans':       safe("SELECT COUNT(*) FROM plans WHERE plan_date=?", today),
            'dos_total':   safe("SELECT COUNT(*) FROM plan_details WHERE plan_date=? AND COALESCE(status,'Planned')!='Cancelled'", today),
            'done':        safe("SELECT COUNT(*) FROM plan_details WHERE plan_date=? AND status='Done'", today),
            'delayed':     safe("SELECT COUNT(*) FROM plan_details WHERE plan_date=? AND status='Delayed'", today),
            'pending':     safe("SELECT COUNT(*) FROM plan_details WHERE plan_date=? AND COALESCE(status,'Planned')='Planned'", today),
            'cancelled':   safe("SELECT COUNT(*) FROM plan_details WHERE plan_date=? AND status='Cancelled'", today),
            'qty_done':    safe("SELECT COALESCE(SUM(do_qty),0) FROM plan_details WHERE plan_date=? AND status IN ('Done','Delayed')", today),
            'qty_total':   safe("SELECT COALESCE(SUM(do_qty),0) FROM plan_details WHERE plan_date=? AND COALESCE(status,'Planned')!='Cancelled'", today),
        },
        'yesterday': {
            'done':        safe("SELECT COUNT(*) FROM plan_details WHERE plan_date=? AND status='Done'", yesterday),
            'pending':     safe("SELECT COUNT(*) FROM plan_details WHERE plan_date=? AND COALESCE(status,'Planned')='Planned'", yesterday),
            'total':       safe("SELECT COUNT(*) FROM plan_details WHERE plan_date=? AND COALESCE(status,'Planned')!='Cancelled'", yesterday),
        },
        'mtd': {
            'plans':       safe("SELECT COUNT(*) FROM plans WHERE plan_date>=?", month_start),
            'dos_done':    safe("SELECT COUNT(*) FROM plan_details WHERE plan_date>=? AND status IN ('Done','Delayed')", month_start),
            'dos_total':   safe("SELECT COUNT(*) FROM plan_details WHERE plan_date>=? AND COALESCE(status,'Planned')!='Cancelled'", month_start),
            'qty_done':    safe("SELECT COALESCE(SUM(do_qty),0) FROM plan_details WHERE plan_date>=? AND status IN ('Done','Delayed')", month_start),
        },
        'wtd': {
            'dos_done':    safe("SELECT COUNT(*) FROM plan_details WHERE plan_date>=? AND status IN ('Done','Delayed')", week_start),
            'dos_total':   safe("SELECT COUNT(*) FROM plan_details WHERE plan_date>=? AND COALESCE(status,'Planned')!='Cancelled'", week_start),
        },
        'total': {
            'plans':       safe("SELECT COUNT(*) FROM plans"),
            'dos':         safe("SELECT COUNT(*) FROM plan_details WHERE COALESCE(status,'Planned')!='Cancelled'"),
            'done':        safe("SELECT COUNT(*) FROM plan_details WHERE status='Done'"),
            'machines':    safe("SELECT COUNT(DISTINCT machine_no) FROM plan_details"),
        },
        'last_7_days': [],
        'scanner_perf': [],
        'floor_summary': [],
        'recent_plans': [],
    }

    # Last 7 days trend
    for i in range(6, -1, -1):
        d = (date.today() - timedelta(days=i)).isoformat()
        total = safe("SELECT COUNT(*) FROM plan_details WHERE plan_date=? AND COALESCE(status,'Planned')!='Cancelled'", d)
        done  = safe("SELECT COUNT(*) FROM plan_details WHERE plan_date=? AND status IN ('Done','Delayed')", d)
        result['last_7_days'].append({'date': d, 'total': total, 'done': done,
            'pct': round(done/max(total,1)*100,1)})

    # Scanner (machine) performance - MTD
    rows = conn.execute("""
        SELECT machine_no, scanner_name,
               COUNT(*) as dos,
               SUM(do_qty) as qty,
               SUM(CASE WHEN status='Done' THEN 1 ELSE 0 END) as done,
               ROUND(AVG(util_pct),1) as avg_util
        FROM plan_details
        WHERE plan_date>=? AND COALESCE(status,'Planned')!='Cancelled'
        GROUP BY machine_no, scanner_name
        ORDER BY dos DESC LIMIT 20
    """, (month_start,)).fetchall()
    result['scanner_perf'] = [dict(r) for r in rows]

    # Floor summary - today
    rows2 = conn.execute("""
        SELECT floor,
               COUNT(*) as total,
               SUM(CASE WHEN status='Done' THEN 1 ELSE 0 END) as done,
               SUM(CASE WHEN COALESCE(status,'Planned')='Planned' THEN 1 ELSE 0 END) as pending,
               SUM(CASE WHEN status='Cancelled' THEN 1 ELSE 0 END) as cancelled,
               SUM(do_qty) as qty
        FROM plan_details WHERE plan_date=?
        GROUP BY floor ORDER BY floor
    """, (today,)).fetchall()
    result['floor_summary'] = [dict(r) for r in rows2]

    # Recent plans
    rows3 = conn.execute("""
        SELECT p.token, p.plan_date, p.run_number, p.total_dos, p.avg_util, p.created_at,
               COUNT(CASE WHEN pd.status='Done' THEN 1 END) as done_count
        FROM plans p LEFT JOIN plan_details pd ON p.token=pd.token
        GROUP BY p.token ORDER BY p.created_at DESC LIMIT 10
    """).fetchall()
    result['recent_plans'] = [dict(r) for r in rows3]

    conn.close()
    return result

# ─── Download DOs by status ────────────────────────────────────────────────────
@app.get("/api/export/dos")
def api_export_dos(token: Optional[str]=None, date_from: Optional[str]=None,
                   date_to: Optional[str]=None, status: Optional[str]=None,
                   floor: Optional[int]=None):
    conn = get_db()
    q = "SELECT pd.*, at.actual_start, at.actual_end, at.actual_date FROM plan_details pd LEFT JOIN actual_times at ON pd.token=at.token AND pd.do_no=at.do_no WHERE 1=1"
    params = []
    if token:     q += " AND pd.token=?";     params.append(token)
    if date_from: q += " AND pd.plan_date>=?"; params.append(date_from)
    if date_to:   q += " AND pd.plan_date<=?"; params.append(date_to)
    if status:    q += " AND COALESCE(pd.status,'Planned')=?"; params.append(status)
    if floor is not None: q += " AND pd.floor=?"; params.append(floor)
    q += " ORDER BY pd.plan_date, pd.floor, pd.priority, pd.do_no"
    rows = conn.execute(q, params).fetchall()
    conn.close()
    if not rows: raise HTTPException(404, "No DOs found matching filters")
    wb = Workbook(); ws = wb.active; ws.title = 'DO_EXPORT'
    def hf(): return Font(name='Arial', size=9, bold=True, color='FFFFFFFF')
    def fl(h): return PatternFill('solid', start_color='FF'+h.lstrip('#'), end_color='FF'+h.lstrip('#'))
    headers = ['TOKEN','PLAN_DATE','FLOOR','PRIORITY','DO_NO','SEC','DO_QTY','PICKER_NO',
               'MACHINE_NO','SCANNER','GRP','BGT','START_TIME','END_TIME','STATUS',
               'CANCEL_REASON','ACTUAL_DATE','ACTUAL_START','ACTUAL_END']
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font = hf(); c.fill = fl('1E3A5F')
        c.alignment = Alignment(horizontal='center')
        ws.column_dimensions[get_column_letter(ci)].width = 14
    STATUS_COLORS_XL = {'Done':'C8F7C5','Delayed':'FDEBD0','Cancelled':'FADBD8',
                        'Not Picked':'D5D8DC','Planned':'D6EAF8'}
    for row in rows:
        r = dict(row)
        st = r.get('status','Planned') or 'Planned'
        bg = STATUS_COLORS_XL.get(st, 'FFFFFF')
        vals = [r.get('token',''), r.get('plan_date',''), r.get('floor',''),
                r.get('priority',''), r.get('do_no',''), r.get('sec',''),
                r.get('do_qty',0), r.get('picker_no',0), r.get('machine_no',''),
                r.get('scanner_name',''), r.get('grp',''), r.get('bgt_machine',0),
                r.get('start_time',''), r.get('end_time',''), st,
                r.get('cancel_reason',''), r.get('actual_date',''),
                r.get('actual_start',''), r.get('actual_end','')]
        ws.append(vals)
        for ci in range(1, len(headers)+1):
            ws.cell(row=ws.max_row, column=ci).fill = fl(bg)
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    fname = f"DO_Export_{date.today().isoformat()}.xlsx"
    return StreamingResponse(io.BytesIO(buf.getvalue()),
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename="{fname}"'})

# ─── Analytics date-range endpoint ────────────────────────────────────────────
@app.get("/api/analytics/summary")
def api_analytics_summary(date_from: Optional[str]=None, date_to: Optional[str]=None):
    conn = get_db()
    q_plans = "SELECT * FROM plans WHERE 1=1"
    q_det   = "SELECT pd.* FROM plan_details pd WHERE 1=1"
    params = []
    if date_from:
        q_plans += " AND plan_date>=?"; q_det += " AND pd.plan_date>=?"; params.append(date_from)
    if date_to:
        q_plans += " AND plan_date<=?"; q_det += " AND pd.plan_date<=?"; params.append(date_to)
    plans = [dict(r) for r in conn.execute(q_plans + " ORDER BY plan_date", params).fetchall()]
    details = [dict(r) for r in conn.execute(q_det, params).fetchall()]
    conn.close()
    if not details: return {'plans': plans, 'details': [], 'summary': {}}
    df = pd.DataFrame(details)
    df['status'] = df['status'].fillna('Planned')
    sc = df['status'].value_counts().to_dict()
    total = len(df)
    return {
        'plans': plans,
        'summary': {
            'total_dos': total, 'total_qty': int(df['do_qty'].sum()),
            'status_counts': sc,
            'completion_pct': round((sc.get('Done',0)+sc.get('Delayed',0))/max(total,1)*100,1),
            'pickers_used': int(df['machine_no'].nunique()),
            'dates': sorted(df['plan_date'].unique().tolist()),
        },
        'by_date': df.groupby('plan_date').apply(lambda g: {
            'dos': len(g), 'qty': int(g['do_qty'].sum()),
            'done': int((g['status']=='Done').sum()),
            'pct': round((g['status'].isin(['Done','Delayed'])).sum()/max(len(g),1)*100,1)
        }).to_dict(),
        'floor_data': df.groupby('floor').apply(lambda g: {
            'total': len(g), 'done': int((g['status']=='Done').sum()),
            'qty': int(g['do_qty'].sum())
        }).to_dict(),
    }

# ─── Email helper ──────────────────────────────────────────────────────────────
def send_email(cfg_dict, to_addr, subject, html_body):
    msg = MIMEMultipart('alternative')
    msg['Subject'] = subject
    msg['From']    = cfg_dict.get('from_addr','')
    msg['To']      = to_addr
    msg.attach(MIMEText(html_body, 'html'))
    s = smtplib.SMTP(cfg_dict['smtp_host'], int(cfg_dict.get('smtp_port',587)))
    s.starttls()
    if cfg_dict.get('smtp_user'):
        s.login(cfg_dict['smtp_user'], cfg_dict['smtp_pass'])
    s.sendmail(cfg_dict['from_addr'], to_addr, msg.as_string())
    s.quit()

# ─── Improved templates with Excel dropdowns ──────────────────────────────────
def make_actuals_template_v2(token, status_filter=None):
    """Actuals template — only required columns + date validation dropdown."""
    details = load_plan_details(token)
    if not details: return None
    if status_filter and status_filter != 'All':
        details = [d for d in details if (d.get('status') or 'Planned') == status_filter]
    wb = Workbook(); ws = wb.active; ws.title = 'ACTUALS'
    from openpyxl.worksheet.datavalidation import DataValidation
    def hf(): return Font(name='Arial', size=9, bold=True, color='FFFFFFFF')
    def fl(h): return PatternFill('solid', start_color='FF'+h.lstrip('#'), end_color='FF'+h.lstrip('#'))
    headers = ['DO_NO','FLOOR','SEC','PRIORITY','DO_QTY','PICKER_NO','MACHINE_NO',
               'PLAN_START','PLAN_END','ACTUAL_DATE','ACTUAL_START','ACTUAL_END','NOTES']
    widths  = [16,7,8,9,10,10,16,12,12,14,13,13,20]
    for ci,(h,w) in enumerate(zip(headers,widths),1):
        c = ws.cell(row=1, column=ci, value=h); c.font = hf(); c.fill = fl('1E3A5F')
        c.alignment = Alignment(horizontal='center')
        ws.column_dimensions[get_column_letter(ci)].width = w
    today_str = date.today().isoformat()
    edit_fill = PatternFill('solid', start_color='FF0A2040', end_color='FF0A2040')
    for i, rec in enumerate(sorted(details, key=lambda r:(r.get('floor',0),r.get('picker_no',0),r.get('start_time','')))):
        r = i + 2
        vals = [str(rec.get('do_no','')), rec.get('floor',0), rec.get('sec',''),
                rec.get('priority',1), rec.get('do_qty',0), rec.get('picker_no',0),
                rec.get('machine_no',''), rec.get('start_time',''), rec.get('end_time',''),
                today_str, '', '', '']
        for ci, v in enumerate(vals, 1):
            cell = ws.cell(row=r, column=ci, value=v)
            if ci >= 10:  # editable cols
                cell.fill = edit_fill
    # Add row 2 hint
    ws.cell(row=2, column=10).comment = None
    buf = io.BytesIO(); wb.save(buf); buf.seek(0); return buf.getvalue()

def make_status_template_v2(token, status_filter=None):
    """Status template with Excel dropdown validation for NEW_STATUS and CANCEL_REASON."""
    details = load_plan_details(token)
    if not details: return None
    if status_filter and status_filter != 'All':
        details = [d for d in details if (d.get('status') or 'Planned') == status_filter]
    wb = Workbook(); ws = wb.active; ws.title = 'STATUS_UPDATE'
    from openpyxl.worksheet.datavalidation import DataValidation
    def hf(): return Font(name='Arial', size=9, bold=True, color='FFFFFFFF')
    def fl(h): return PatternFill('solid', start_color='FF'+h.lstrip('#'), end_color='FF'+h.lstrip('#'))
    headers = ['DO_NO','FLOOR','SEC','PRIORITY','DO_QTY','PICKER_NO','MACHINE_NO',
               'PLAN_START','PLAN_END','CURRENT_STATUS','NEW_STATUS','CANCEL_REASON']
    widths  = [16,7,8,9,10,10,16,12,12,14,14,25]
    for ci,(h,w) in enumerate(zip(headers,widths),1):
        c = ws.cell(row=1, column=ci, value=h); c.font = hf(); c.fill = fl('1E3A5F')
        c.alignment = Alignment(horizontal='center')
        ws.column_dimensions[get_column_letter(ci)].width = w
    # Excel dropdown validations
    dv_status = DataValidation(type='list', formula1='"Done,Cancelled,Delayed,Not Picked,Planned"', allow_blank=True)
    dv_reason = DataValidation(type='list', formula1='"Picker absent,Stock unavailable,DO cancelled by WH,End of shift,System error,Other"', allow_blank=True)
    ws.add_data_validation(dv_status); ws.add_data_validation(dv_reason)
    edit_fill = PatternFill('solid', start_color='FF0A2040', end_color='FF0A2040')
    for i, rec in enumerate(sorted(details, key=lambda r:(r.get('floor',0),r.get('picker_no',0),r.get('start_time','')))):
        r = i + 2
        vals = [str(rec.get('do_no','')), rec.get('floor',0), rec.get('sec',''),
                rec.get('priority',1), rec.get('do_qty',0), rec.get('picker_no',0),
                rec.get('machine_no',''), rec.get('start_time',''), rec.get('end_time',''),
                rec.get('status','Planned'), '', '']
        for ci, v in enumerate(vals, 1):
            cell = ws.cell(row=r, column=ci, value=v)
            if ci >= 11: cell.fill = edit_fill
        dv_status.add(ws.cell(row=r, column=11))
        dv_reason.add(ws.cell(row=r, column=12))
    buf = io.BytesIO(); wb.save(buf); buf.seek(0); return buf.getvalue()

# ─── Updated template endpoints with filter support ───────────────────────────
@app.get("/api/plans/{token}/actuals-template-v2")
def api_actuals_template_v2(token: str, status_filter: Optional[str]=None):
    xlsx = make_actuals_template_v2(token, status_filter)
    if not xlsx: raise HTTPException(404, "No details for token")
    return StreamingResponse(io.BytesIO(xlsx),
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename="Actuals_{token}.xlsx"'})

@app.get("/api/plans/{token}/status-template-v2")
def api_status_template_v2(token: str, status_filter: Optional[str]=None):
    xlsx = make_status_template_v2(token, status_filter)
    if not xlsx: raise HTTPException(404, "No details for token")
    return StreamingResponse(io.BytesIO(xlsx),
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename="Status_{token}.xlsx"'})

# ─── Serve React Frontend ─────────────────────────────────────────────────────
STATIC_DIR = os.path.join(BASE_DIR, "static")
if os.path.exists(STATIC_DIR):
    app.mount("/static", StaticFiles(directory=STATIC_DIR), name="static")

@app.get("/{full_path:path}")
def serve_spa(full_path: str):
    index_path = os.path.join(STATIC_DIR, "index.html")
    if os.path.exists(index_path):
        from fastapi.responses import FileResponse
        return FileResponse(index_path)
    return {"message": "Picker Planning API", "docs": "/docs"}
