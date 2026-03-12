"""Excel export utilities."""
import io
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

P_COLS = ['#E74C3C','#E67E22','#F1C40F','#2ECC71','#1ABC9C','#3498DB',
          '#9B59B6','#E91E63','#FF5722','#795548','#607D8B','#00BCD4','#8BC34A','#FFC107']
GRP_CLR = {'G1':'#2ECC71','G2':'#4F8EF7','G3':'#F39C12'}

def hf(sz=10,bd=False,cl='FFFFFF'): return Font(name='DM Sans',size=sz,bold=bd,color=cl)
def fl(c): return PatternFill('solid',fgColor=c)
def al(h='center',v='center',w=True): return Alignment(horizontal=h,vertical=v,wrap_text=w)

def make_plan_excel(plan_df, cfg, meta=None):
    BGT = cfg.get('bgt',3000); FILL = cfg.get('fill_pct',70)
    token = (meta or {}).get('token','')
    plan_date = (meta or {}).get('plan_date','')
    run_num = (meta or {}).get('run_number','')

    wb = Workbook(); wb.remove(wb.active)
    def hf(c='FFFFFFFF',sz=10): return Font(name='Arial',size=sz,bold=True,color=c)
    def cf(sz=9,b=False,c='FF000000'): return Font(name='Arial',size=sz,bold=b,color=c)
    def fl(h):
        h=h.lstrip('#'); h=('FF'+h) if len(h)==6 else h
        return PatternFill('solid',start_color=h,end_color=h)
    def bd():
        s=Side(style='thin',color='FFBDC3C7')
        return Border(left=s,right=s,top=s,bottom=s)
    def al(h='center'): return Alignment(horizontal=h,vertical='center',wrap_text=False)
    PC_X={1:('FDEDEC','C0392B'),2:('FEF9E7','F39C12'),3:('EAF4FB','2874A6'),4:('EAFAF1','1E8449')}
    GCLR={'G1':'FF1E8449','G2':'FF2874A6','G3':'FFF39C12'}

    # Normalise col names from DB
    df = plan_df.copy()
    CMAP={'grp':'GROUP','do_qty':'DO_QTY','picker_no':'PICKER_NO','machine_no':'MACHINE_NO',
          'scanner_name':'SCANNER_NAME','bgt_machine':'BGT_MACHINE','start_time':'START_TIME',
          'end_time':'END_TIME','duration_min':'DURATION_MIN','pcs_per_min':'PCS_PER_MIN',
          'cap_used':'CAP_USED','util_pct':'UTIL_%','remaining':'REMAINING','over_wh':'OVER_WH',
          'floor':'FLOOR','sec':'SEC','do_no':'DO_NO','sto_no':'STO_NO','st_cd':'ST_CD',
          'st_nm':'ST_NM','priority':'PRIORITY'}
    df.rename(columns={k:v for k,v in CMAP.items() if k in df.columns}, inplace=True)

    # SUMMARY
    ss=wb.create_sheet('SUMMARY',0)
    ss.merge_cells('A1:F1')
    ss['A1']=f'📊  PICKER PLAN SUMMARY   {token}   {plan_date}   Run #{run_num}'
    ss['A1'].font=hf(sz=12); ss['A1'].fill=fl('1C2833'); ss['A1'].alignment=al(); ss.row_dimensions[1].height=26
    g1c=int((df['GROUP']=='G1').sum()) if 'GROUP' in df.columns else 0
    g2c=int((df['GROUP']=='G2').sum()) if 'GROUP' in df.columns else 0
    g3c=int((df['GROUP']=='G3').sum()) if 'GROUP' in df.columns else 0
    for i,((lbl,val),bg) in enumerate(zip([
        ('Token',token),('Plan Date',plan_date),('Run #',str(run_num)),
        ('Total DOs',f"{len(df):,}"),
        ('Total QTY',f"{int(df['DO_QTY'].sum()):,}" if 'DO_QTY' in df.columns else '0'),
        ('Fill %',f"{FILL}%")],
        ['1A5276','117A65','6E2FD3','922B21','1F618D','7D6608']),1):
        ss.column_dimensions[get_column_letter(i)].width=18
        lc=ss.cell(row=3,column=i,value=lbl); lc.font=hf(sz=8); lc.fill=fl(bg)
        lc.alignment=Alignment(horizontal='center',vertical='bottom')
        vc=ss.cell(row=4,column=i,value=str(val))
        vc.font=Font(name='Arial',size=13,bold=True,color='FF'+bg)
        vc.fill=fl('FFFFFF'); vc.alignment=al()
    ss.row_dimensions[3].height=30; ss.row_dimensions[4].height=24
    ss.sheet_view.showGridLines=False

    # PICKER_PLAN
    ws=wb.create_sheet('PICKER_PLAN')
    H=['PRIORITY','DO_NO','STO_NO','ST_CD','ST_NM','FLOOR','SEC','DO_QTY',
       'PICKER_NO','MACHINE_NO','SCANNER_NAME','GROUP','BGT_MACHINE',
       'START_TIME','END_TIME','DURATION_MIN','PCS_PER_MIN','CAP_USED','UTIL_%','REMAINING']
    H=[c for c in H if c in df.columns]; NC=len(H)
    ws.merge_cells(f'A1:{get_column_letter(NC)}1')
    ws['A1']=f'🏭  PICKER PLAN   {token}   G1/G2/G3 Group-Aware · Continuous Greedy · No-Split'
    ws['A1'].font=hf(sz=11); ws['A1'].fill=fl('1C2833'); ws['A1'].alignment=al(); ws.row_dimensions[1].height=24
    W=[10,14,14,10,26,7,7,10,10,14,16,8,12,12,12,13,13,12,11,12]
    for c,(h,w) in enumerate(zip(H,W[:len(H)]),1):
        cell=ws.cell(row=2,column=c,value=h)
        cell.font=hf(); cell.fill=fl('2C3E50'); cell.alignment=al(); cell.border=bd()
        ws.column_dimensions[get_column_letter(c)].width=w
    ws.row_dimensions[2].height=20
    prev_p=None
    for i,rec in df.iterrows():
        r=3+i; p=int(rec.get('PRIORITY',0)); over=bool(rec.get('OVER_WH',0))
        pbg,ptxt=PC_X.get(p,('F8F9FA','7F8C8D'))
        rf='FFF5E6' if over else ('F0F8FF' if i%2==0 else 'FFFFFF')
        for c,fld in enumerate(H,1):
            cell=ws.cell(row=r,column=c,value=rec.get(fld,'')); cell.border=bd()
            cell.fill=fl(rf); cell.font=cf(); cell.alignment=al('left') if c in(5,11) else al()
        ws.cell(row=r,column=1).fill=fl(pbg)
        ws.cell(row=r,column=1).font=Font(name='Arial',size=9,bold=True,color='FF'+ptxt)
        if 'GROUP' in H:
            gi=H.index('GROUP')+1; grpv=str(rec.get('GROUP','G1'))
            ws.cell(row=r,column=gi).font=Font(name='Arial',size=9,bold=True,color=GCLR.get(grpv,'FF000000'))
        if 'UTIL_%' in H:
            ui=H.index('UTIL_%')+1; u=float(rec.get('UTIL_%',0))
            ws.cell(row=r,column=ui).font=Font(name='Arial',size=9,bold=True,
                color='FF1E8449' if u>=FILL else('FFF39C12' if u>=50 else 'FFC0392B'))
        if over and 'END_TIME' in H:
            ei=H.index('END_TIME')+1
            ws.cell(row=r,column=ei).font=Font(name='Arial',size=9,bold=True,color='FFC0392B')
        if prev_p and p!=prev_p:
            for c in range(1,NC+1):
                ws.cell(row=r,column=c).border=Border(
                    top=Side(style='medium',color='FF2C3E50'),
                    left=Side(style='thin',color='FFBDC3C7'),
                    right=Side(style='thin',color='FFBDC3C7'),
                    bottom=Side(style='thin',color='FFBDC3C7'))
        prev_p=p; ws.row_dimensions[r].height=15
    ws.freeze_panes='A3'; ws.sheet_view.showGridLines=False
    if len(df): ws.auto_filter.ref=f'A2:{get_column_letter(NC)}{3+len(df)-1}'

    # PICKER_SCHEDULE
    pf=wb.create_sheet('PICKER_SCHEDULE')
    pf.merge_cells('A1:L1')
    pf['A1']='👷  PICKER FULL-DAY SCHEDULE'
    pf['A1'].font=hf(sz=12); pf['A1'].fill=fl('1C2833'); pf['A1'].alignment=al(); pf.row_dimensions[1].height=24
    SCH_H=['FLOOR','PICKER#','MACHINE_NO','SCANNER_NAME','GROUP','BGT_MACHINE',
            'DOs','QTY_ASSIGNED','PRIORITIES','DAY_START','DAY_END','FINAL_UTIL_%']
    for c,(h,w) in enumerate(zip(SCH_H,[7,9,14,18,8,12,8,14,35,11,11,12]),1):
        cell=pf.cell(row=2,column=c,value=h)
        cell.font=hf(); cell.fill=fl('2C3E50'); cell.alignment=al(); cell.border=bd()
        pf.column_dimensions[get_column_letter(c)].width=w
    pf.row_dimensions[2].height=20
    sc_g=['FLOOR','PICKER_NO','MACHINE_NO','SCANNER_NAME']
    if 'GROUP' in df.columns: sc_g.append('GROUP')
    if 'BGT_MACHINE' in df.columns: sc_g.append('BGT_MACHINE')
    sc_g=[c for c in sc_g if c in df.columns]
    sched=(df.groupby(sc_g)
           .agg(DOs=('DO_NO','count'),QTY=('DO_QTY','sum'),
                PRIOS=('PRIORITY',lambda x:' → '.join([f'P{int(p)}' for p in sorted(x.unique())])),
                START=('START_TIME','min'),END=('END_TIME','max'),UTIL=('UTIL_%','max'))
           .reset_index().sort_values(['FLOOR','PICKER_NO']))
    wh_str=cfg.get('wh_end_str','17:00')
    for i,row in sched.iterrows():
        r=3+i; u=float(row['UTIL']); over=str(row['END'])>wh_str
        grpv=str(row.get('GROUP','G1')) if 'GROUP' in row.index else 'G1'
        bgtv=int(row.get('BGT_MACHINE',BGT)) if 'BGT_MACHINE' in row.index else BGT
        gc=GCLR.get(grpv,'FF000000')
        rf='FFF5E6' if over else ('EAFAF1' if u>=FILL else ('FEF9E7' if u>=50 else 'FDEDEC'))
        pf.row_dimensions[r].height=17
        vals=[int(row['FLOOR']),int(row['PICKER_NO']),row['MACHINE_NO'],row['SCANNER_NAME'],
              grpv,bgtv,int(row['DOs']),int(row['QTY']),row['PRIOS'],row['START'],row['END'],round(u,1)]
        for c,v in enumerate(vals,1):
            cell=pf.cell(row=r,column=c,value=v); cell.font=cf(sz=10)
            cell.alignment=al(); cell.border=bd(); cell.fill=fl(rf)
            if c==5: cell.font=Font(name='Arial',size=10,bold=True,color=gc)
            if c==12: cell.font=Font(name='Arial',size=10,bold=True,
                color='FF1E8449' if u>=FILL else('FFF39C12' if u>=50 else 'FFC0392B'))
            if c==11 and over: cell.font=Font(name='Arial',size=10,bold=True,color='FFC0392B')
    pf.freeze_panes='A3'; pf.sheet_view.showGridLines=False

    # ACTUAL TIME ENTRY template
    at=wb.create_sheet('ACTUAL_TIME_ENTRY')
    at.merge_cells('A1:I1')
    at['A1']='⏱  ACTUAL TIME ENTRY TEMPLATE — fill ACTUAL_START and ACTUAL_END (HH:MM format)'
    at['A1'].font=hf(sz=10); at['A1'].fill=fl('1C2833'); at['A1'].alignment=al(); at.row_dimensions[1].height=22


def make_actuals_template(details, plan_meta):
    """Build actuals entry template Excel."""
    wb = Workbook(); ws = wb.active; ws.title = "ACTUALS"
    cols = ['DO_NO','FLOOR','SEC','PRIORITY','DO_QTY','PICKER_NO',
            'PLAN_START','PLAN_END','ACTUAL_DATE','ACTUAL_START','ACTUAL_END','NOTES']
    hdr_fill = fl('1C2836')
    for ci,col in enumerate(cols,1):
        c = ws.cell(row=1,column=ci,value=col)
        c.font=hf(9,True,'4F8EF7'); c.fill=hdr_fill; c.alignment=al()
        ws.column_dimensions[get_column_letter(ci)].width=15
    plan_date = plan_meta.get('plan_date','')
    ef = fl('0A2040')
    for _,r in details.sort_values(['floor','picker_no','start_time']).iterrows():
        ws.append([str(r.get('do_no','')), int(r.get('floor',0)),
                   str(r.get('sec','')), int(r.get('priority',1)),
                   int(r.get('do_qty',0)), int(r.get('picker_no',0)),
                   str(r.get('start_time','')), str(r.get('end_time','')),
                   plan_date,'','',''])
        for ci in [9,10,11,12]:
            ws.cell(row=ws.max_row,column=ci).fill=ef
    buf=io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf.getvalue()


def make_status_template(details, plan_meta):
    """Build status update template Excel."""
    wb = Workbook(); ws = wb.active; ws.title = "STATUS_UPDATE"
    cols = ['DO_NO','FLOOR','SEC','PRIORITY','DO_QTY','PICKER_NO',
            'PLAN_START','PLAN_END','CURRENT_STATUS','NEW_STATUS','CANCEL_REASON']
    hdr_fill = fl('1C2836')
    for ci,col in enumerate(cols,1):
        c = ws.cell(row=1,column=ci,value=col)
        c.font=hf(9,True,'4F8EF7'); c.fill=hdr_fill; c.alignment=al()
        ws.column_dimensions[get_column_letter(ci)].width=17
    ef = fl('0A2040')
    for _,r in details.sort_values(['floor','picker_no','start_time']).iterrows():
        ws.append([str(r.get('do_no','')), int(r.get('floor',0)),
                   str(r.get('sec','')), int(r.get('priority',1)),
                   int(r.get('do_qty',0)), int(r.get('picker_no',0)),
                   str(r.get('start_time','')), str(r.get('end_time','')),
                   str(r.get('status','Planned')),'',''])
        for ci in [10,11]:
            ws.cell(row=ws.max_row,column=ci).fill=ef
    buf=io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf.getvalue()
